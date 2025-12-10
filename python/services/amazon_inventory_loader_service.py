#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Amazon Inventory Loader (Lファイル) 生成サービス

Amazon出品用のInventory Loader TSVファイルを生成する。
画像URLを含む出品用ファイルフォーマットに対応。
"""
from __future__ import annotations

import pandas as pd
import io
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)

# コンディションマッピング（Amazonコンディション番号）
CONDITION_MAP = {
    "中古(ほぼ新品)": "1",
    "中古(非常に良い)": "2", 
    "中古(良い)": "3",
    "中古(可)": "4",
    "コレクター商品(ほぼ新品)": "5",
    "コレクター商品(非常に良い)": "6",
    "コレクター商品(良い)": "7",
    "コレクター商品(可)": "8",
    "再生品": "10",
    "新品(新品)": "11",
    "新品": "11",
}


class AmazonInventoryLoaderService:
    """Amazon Inventory Loader (Lファイル) 生成サービス"""
    
    @staticmethod
    def generate_inventory_loader_tsv(products: List[dict]) -> bytes:
        """
        Amazon Inventory Loader TSVファイルを生成（.txt形式、タブ区切りテキスト）
        
        理由: 自作のExcelファイル(.xlsx)はAmazonシステムに「サポート外」として拒否されるため。
        Pythonで厳密に列定義されたTSV(タブ区切り)を生成することで、列ズレを防ぐ。
        
        Args:
            products: 商品データのリスト
                必須フィールド:
                - sku: SKU
                - product_id: ASIN または JAN
                - product_id_type: 1=ASIN, 4=JAN
                - item_condition: コンディション番号（11=新品, 1=ほぼ新品, 2=非常に良い...）
                オプションフィールド:
                - price: 販売価格（空欄可）
                - quantity: 在庫数（空欄可、FBA指定のため）
                - add_delete: 登録/削除フラグ（デフォルト: 'a' = add/update）
                - image_urls: 商品画像URLのリスト（最大6枚まで、GCSアップロード後）
                  または image_url_1 ～ image_url_6: 個別の画像URLキー
                
        Returns:
            TSV bytes (.txt形式、cp932/Shift-JIS、タブ区切り)
        """
        if not products:
            return b""
        
        # ヘッダー定義（Amazon Inventory Loader形式 - 出品ファイル(I)形式）
        # 【TSV形式対応版】Inventory Loaderの標準カラム順序（Amazonテンプレートに準拠）
        # 【重要】列の並び順を固定することで、列ズレを防ぎます。
        required_columns = [
            'sku',                              # 1. 商品管理番号
            'product-id',                       # 2. ASIN または JAN
            'product-id-type',                  # 3. IDタイプ (1=ASIN, 4=JAN)
            'price',                            # 4. 価格
            'minimum-seller-allowed-price',     # 5. 最低販売価格
            'maximum-seller-allowed-price',     # 6. 最高販売価格
            'item-condition',                   # 7. コンディション (11=新品, 1=ほぼ新品...)
            'quantity',                         # 8. 在庫数 (FBA納品前は空欄)
            'add-delete',                       # 9. 登録/削除 (a=追加/更新)
            'will-ship-internationally',        # 10. 国際配送
            'expedited-shipping',               # 11. 速達配送
            'standard-plus',                    # 12. 標準プラス
            'item-note',                        # 13. 商品メモ
            'fulfillment-center-id',            # 14. FBA指定 (AMAZON_JP)
            # 画像カラム群（最大6枚まで対応）
            'main-offer-image-url',             # 15. コンディション写真（メイン・1枚目）
            'offer-image1',                     # 16. コンディション写真（サブ1・2枚目）
            'offer-image2',                     # 17. コンディション写真（サブ2・3枚目）
            'offer-image3',                     # 18. コンディション写真（サブ3・4枚目）
            'offer-image4',                     # 19. コンディション写真（サブ4・5枚目）
            'offer-image5',                     # 20. コンディション写真（サブ5・6枚目）
        ]
        
        # データ変換
        records = []
        for product in products:
            # ===== 必須フィールドのチェック =====
            sku = product.get('sku') or product.get('SKU')
            if not sku:
                logger.warning(f"Skipping product without SKU: {product}")
                continue

            # A案: 画像だけ更新モード
            # 価格・在庫は常に空文字列を送る（Amazon側で価格・在庫を更新しない想定）
            # TSV形式では空文字列が空欄として出力される
            price_value: Any = ''
            quantity_value: Any = ''
            
            # product-id と product-id-type
            asin = product.get('asin') or product.get('ASIN') or ''
            jan = product.get('jan') or product.get('JAN') or ''
            
            if asin:
                product_id = asin
                product_id_type = 1  # ASIN
            elif jan:
                product_id = jan
                product_id_type = 4  # JAN
            else:
                logger.warning(f"Skipping product {sku} without ASIN or JAN")
                continue
            
            # コンディション（item-conditionに変更）
            item_condition = product.get('item_condition') or product.get('item-condition') or product.get('condition_type') or product.get('condition-type')
            if not item_condition:
                # コンディション文字列から変換
                condition_str = product.get('condition') or product.get('コンディション') or ''
                item_condition = CONDITION_MAP.get(condition_str, '')
            
            if not item_condition:
                logger.warning(f"Skipping product {sku} without condition")
                continue
            
            # 画像URLの取得（最大6枚まで対応）
            # image_urlsリスト形式またはimage_url_1～6の個別キー形式の両方に対応
            image_urls = []
            
            # 1. image_urlsリスト形式を優先的に確認
            if 'image_urls' in product and isinstance(product.get('image_urls'), list):
                image_urls = [url for url in product['image_urls'] if url]
            else:
                # 2. image_url_1～6の個別キー形式を確認
                for i in range(1, 7):  # 6枚まで取得
                    img_key = f'image_url_{i}'
                    img_url = product.get(img_key) or product.get(f'画像URL{i}') or ''
                    if img_url:
                        image_urls.append(img_url)
            
            # 最大6枚まで（main-offer-image-url + offer-image1～5）
            # レコード作成（ヘッダーの順番通りに）
            record = {
                'sku': str(sku),
                'product-id': str(product_id),
                'product-id-type': int(product_id_type),
                # A案: 価格・在庫は空文字列を許容（画像だけ更新モード）
                'price': price_value,
                'minimum-seller-allowed-price': '',
                'maximum-seller-allowed-price': '',
                'item-condition': int(item_condition) if item_condition else '',
                'quantity': quantity_value,
                'add-delete': 'a',  # a = add/update
                'will-ship-internationally': '',
                'expedited-shipping': '',
                'standard-plus': '',
                'item-note': '',
                'fulfillment-center-id': 'AMAZON_JP',  # FBA利用
                # 画像URLの設定（最大6枚まで）
                'main-offer-image-url': image_urls[0] if len(image_urls) > 0 else '',  # 1枚目: メイン出品画像
                'offer-image1': image_urls[1] if len(image_urls) > 1 else '',          # 2枚目
                'offer-image2': image_urls[2] if len(image_urls) > 2 else '',          # 3枚目
                'offer-image3': image_urls[3] if len(image_urls) > 3 else '',          # 4枚目
                'offer-image4': image_urls[4] if len(image_urls) > 4 else '',          # 5枚目
                'offer-image5': image_urls[5] if len(image_urls) > 5 else '',           # 6枚目
            }
            records.append(record)
        
        if not records:
            return b""
        
        # DataFrameに変換
        df = pd.DataFrame(records)
        
        # データに無いカラムは空文字列で埋める
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
        
        # カラム並び替え（これが最強のズレ防止策です）
        df = df[required_columns]
        
        # TSV形式 (.txt) で出力
        # sep='\t' でタブ区切りを指定
        # encoding='cp932' (Shift-JIS) はWindowsのExcelで文字化けせず開ける形式です
        output = io.StringIO()
        try:
            df.to_csv(output, sep='\t', index=False, encoding='cp932', errors='ignore')
            tsv_str = output.getvalue()
            tsv_bytes = tsv_str.encode('cp932', errors='ignore')
            logger.info(f"Generated Amazon Inventory Loader TSV: {len(records)} records")
            return tsv_bytes
        except Exception as e:
            logger.error(f"Failed to generate TSV file: {e}")
            raise
        finally:
            output.close()
    
    @staticmethod
    def generate_listing_loader_tsv(products: List[dict]) -> bytes:
        """
        Amazon Listing Loader（出品ファイルL）TSVファイルを生成（.txt形式、タブ区切りテキスト）
        
        画像登録のために必要な最小限のカラムのみを使用します。
        price, quantity, expedited-shipping など、不要な列は一切含めません。
        
        Args:
            products: 商品データのリスト
                必須フィールド:
                - sku: SKU
                - product_id: ASIN または JAN
                - product_id_type: 1=ASIN, 4=JAN
                - condition_type: コンディション番号（11=新品, 1=ほぼ新品, 2=非常に良い...）
                オプションフィールド:
                - condition_note: コンディション説明（空欄可）
                - operation_type: 操作タイプ（デフォルト: 'Update'）
                - image_urls: 商品画像URLのリスト（最大6枚まで、GCSアップロード後）
                  または image_url_1 ～ image_url_6: 個別の画像URLキー
                
        Returns:
            TSV bytes (.txt形式、cp932/Shift-JIS、タブ区切り)
        """
        if not products:
            return b""
        
        # ヘッダー定義（Amazon Listing Loader形式 - 出品ファイル(L)形式）
        # 【重要】画像登録のための最小必須カラム順序
        # price, quantity, expedited-shipping など、不要な列は一切含めません。
        required_columns = [
            'sku',                              # 1. 商品管理番号
            'product-id',                       # 2. ASIN または JAN
            'product-id-type',                  # 3. IDタイプ (1=ASIN, 4=JAN)
            'condition-type',                   # 4. コンディション (11=新品, 1=ほぼ新品...)
            'condition-note',                    # 5. コンディション説明（空欄でOK）
            'operation-type',                    # 6. 操作タイプ (Update = 更新)
            'fulfillment-center-id',            # 7. FBA指定 (AMAZON_JP)
            # 画像カラム群（最大6枚まで対応）
            'main-offer-image',                 # 8. コンディション写真（メイン・1枚目）
            'offer-image1',                     # 9. コンディション写真（サブ1・2枚目）
            'offer-image2',                     # 10. コンディション写真（サブ2・3枚目）
            'offer-image3',                     # 11. コンディション写真（サブ3・4枚目）
            'offer-image4',                     # 12. コンディション写真（サブ4・5枚目）
            'offer-image5',                     # 13. コンディション写真（サブ5・6枚目）
        ]
        
        # データ変換
        records = []
        for product in products:
            # ===== 必須フィールドのチェック =====
            sku = product.get('sku') or product.get('SKU')
            if not sku:
                logger.warning(f"Skipping product without SKU: {product}")
                continue
            
            # product-id と product-id-type
            asin = product.get('asin') or product.get('ASIN') or ''
            jan = product.get('jan') or product.get('JAN') or ''
            
            if asin:
                product_id = asin
                product_id_type = 1  # ASIN
            elif jan:
                product_id = jan
                product_id_type = 4  # JAN
            else:
                logger.warning(f"Skipping product {sku} without ASIN or JAN")
                continue
            
            # コンディション（condition-type）
            condition_type = product.get('condition_type') or product.get('condition-type') or product.get('item_condition') or product.get('item-condition')
            if not condition_type:
                # コンディション文字列から変換
                condition_str = product.get('condition') or product.get('コンディション') or ''
                condition_type = CONDITION_MAP.get(condition_str, '')
            
            if not condition_type:
                logger.warning(f"Skipping product {sku} without condition")
                continue
            
            # コンディション説明（condition-note）
            condition_note = product.get('condition_note') or product.get('condition-note') or ''
            
            # 操作タイプ（operation-type）
            operation_type = product.get('operation_type') or product.get('operation-type') or 'Update'
            
            # 画像URLの取得（最大6枚まで対応）
            # image_urlsリスト形式またはimage_url_1～6の個別キー形式の両方に対応
            image_urls = []
            
            # 1. image_urlsリスト形式を優先的に確認
            if 'image_urls' in product and isinstance(product.get('image_urls'), list):
                image_urls = [url for url in product['image_urls'] if url]
            else:
                # 2. image_url_1～6の個別キー形式を確認
                for i in range(1, 7):  # 6枚まで取得
                    img_key = f'image_url_{i}'
                    img_url = product.get(img_key) or product.get(f'画像URL{i}') or ''
                    if img_url:
                        image_urls.append(img_url)
            
            # 最大6枚まで（main-offer-image + offer-image1～5）
            # レコード作成（ヘッダーの順番通りに）
            record = {
                'sku': str(sku),
                'product-id': str(product_id),
                'product-id-type': int(product_id_type),
                'condition-type': int(condition_type) if condition_type else '',
                'condition-note': str(condition_note) if condition_note else '',
                'operation-type': str(operation_type),
                'fulfillment-center-id': 'AMAZON_JP',  # FBA利用
                # 画像URLの設定（最大6枚まで）
                'main-offer-image': image_urls[0] if len(image_urls) > 0 else '',  # 1枚目: メイン出品画像
                'offer-image1': image_urls[1] if len(image_urls) > 1 else '',      # 2枚目
                'offer-image2': image_urls[2] if len(image_urls) > 2 else '',      # 3枚目
                'offer-image3': image_urls[3] if len(image_urls) > 3 else '',      # 4枚目
                'offer-image4': image_urls[4] if len(image_urls) > 4 else '',      # 5枚目
                'offer-image5': image_urls[5] if len(image_urls) > 5 else '',      # 6枚目
            }
            records.append(record)
        
        if not records:
            return b""
        
        # DataFrameに変換
        df = pd.DataFrame(records)
        
        # データに無いカラムは空文字列で埋める
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''
        
        # カラム並び替え（これが最強のズレ防止策です）
        df = df[required_columns]
        
        # TSV形式 (.txt) で出力
        # sep='\t' でタブ区切りを指定
        # encoding='cp932' (Shift-JIS) はWindowsのExcelで文字化けせず開ける形式です
        output = io.StringIO()
        try:
            df.to_csv(output, sep='\t', index=False, encoding='cp932', errors='ignore')
            tsv_str = output.getvalue()
            tsv_bytes = tsv_str.encode('cp932', errors='ignore')
            logger.info(f"Generated Amazon Listing Loader TSV: {len(records)} records")
            return tsv_bytes
        except Exception as e:
            logger.error(f"Failed to generate Listing Loader TSV file: {e}")
            raise
        finally:
            output.close()
    
    @staticmethod
    def generate_inventory_loader_excel(products: List[dict]) -> bytes:
        """
        Amazon Inventory Loader Excelファイルを生成（後方互換性のため残す）
        
        【非推奨】TSV形式（generate_inventory_loader_tsv）の使用を推奨します。
        自作のExcelファイル(.xlsx)はAmazonシステムに「サポート外」として拒否されるため。
        
        Args:
            products: 商品データのリスト
            
        Returns:
            Excel bytes (.xlsx形式)
        """
        # 後方互換性のため、TSV形式を生成して返す
        logger.warning("Excel format is deprecated. Use TSV format (generate_inventory_loader_tsv) instead.")
        return AmazonInventoryLoaderService.generate_inventory_loader_tsv(products)
    
    @staticmethod
    def read_amazon_template_excel(template_path: str) -> dict:
        """
        AmazonテンプレートExcelファイルの「テンプレート」シートを読み込む
        
        Args:
            template_path: AmazonテンプレートExcelファイルのパス
            
        Returns:
            テンプレート情報の辞書:
            - headers: ヘッダー行のリスト（列番号とヘッダー名のタプル）
            - max_row: 最大行数
            - max_column: 最大列数
            - sheet_names: シート名のリスト
        """
        try:
            import openpyxl
            from pathlib import Path
            
            file_path = Path(template_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            wb = openpyxl.load_workbook(file_path, keep_vba=True)
            
            result = {
                "sheet_names": wb.sheetnames,
                "template_sheet": None
            }
            
            if "テンプレート" in wb.sheetnames:
                ws = wb["テンプレート"]
                headers = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    if cell_value:
                        headers.append((col_idx, str(cell_value)))
                
                result["template_sheet"] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "headers": headers
                }
            else:
                logger.warning(f"'テンプレート' sheet not found in {template_path}. Available sheets: {wb.sheetnames}")
            
            return result
            
        except ImportError:
            raise ImportError("openpyxl is required. Please install it with: pip install openpyxl")
        except Exception as e:
            logger.error(f"Failed to read Amazon template Excel: {e}")
            raise
    
    @staticmethod
    def write_to_amazon_template_excel(
        template_path: str,
        products: List[dict],
        output_path: Optional[str] = None,
        start_row: int = 7
    ) -> str:
        """
        AmazonテンプレートExcelファイルに商品データを書き込む
        
        Args:
            template_path: AmazonテンプレートExcelファイルのパス
            products: 商品データのリスト
                必須フィールド:
                - sku: SKU
                オプションフィールド:
                - image_urls: 商品画像URLのリスト（最大6枚まで）
                  または image_url_1 ～ image_url_6: 個別の画像URLキー
                - asin: ASIN（必要に応じて）
            output_path: 出力先ファイルパス（Noneの場合はテンプレートファイルを上書き）
            start_row: データ書き込み開始行（デフォルト: 7）
            
        Returns:
            出力先ファイルパス
        """
        try:
            import openpyxl
            from pathlib import Path
            from openpyxl.utils import get_column_letter
            
            file_path = Path(template_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            # テンプレートファイルを読み込む
            wb = openpyxl.load_workbook(file_path, keep_vba=True)
            
            # 「テンプレート」シートを取得
            if "テンプレート" not in wb.sheetnames:
                raise ValueError(f"'テンプレート' sheet not found in {template_path}. Available sheets: {wb.sheetnames}")
            
            ws = wb["テンプレート"]
            
            # データ書き込み位置の定義
            # SKU: A列 (column=1)
            # 画像URL: O列(15)からT列(20)まで
            sku_col = 1  # A列
            image_cols = [15, 16, 17, 18, 19, 20]  # O, P, Q, R, S, T列
            
            # 商品データを書き込む
            for idx, product in enumerate(products):
                row = start_row + idx
                
                # SKUを書き込み
                sku = product.get('sku') or product.get('SKU')
                if sku:
                    ws.cell(row=row, column=sku_col, value=str(sku))
                
                # 画像URLを取得（最大6枚まで）
                image_urls = []
                
                # 1. image_urlsリスト形式を優先的に確認
                if 'image_urls' in product and isinstance(product.get('image_urls'), list):
                    image_urls = [url for url in product['image_urls'] if url]
                else:
                    # 2. image_url_1～6の個別キー形式を確認
                    for i in range(1, 7):  # 6枚まで取得
                        img_key = f'image_url_{i}'
                        img_url = product.get(img_key) or product.get(f'画像URL{i}') or ''
                        if img_url:
                            image_urls.append(img_url)
                
                # 画像URLを書き込み（最大6枚まで）
                for img_idx, img_url in enumerate(image_urls[:6]):
                    if img_idx < len(image_cols):
                        ws.cell(row=row, column=image_cols[img_idx], value=str(img_url))
            
            # 出力先ファイルパスを決定
            if output_path is None:
                # テンプレートファイルと同じディレクトリに、タイムスタンプ付きで保存
                output_dir = file_path.parent
                timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                output_path = str(output_dir / f"ListingLoader_{timestamp}.xlsm")
            else:
                output_path = str(Path(output_path))
            
            # ファイルを保存
            wb.save(output_path)
            logger.info(f"Written {len(products)} products to Amazon template Excel: {output_path}")
            
            return output_path
            
        except ImportError:
            raise ImportError("openpyxl is required. Please install it with: pip install openpyxl")
        except Exception as e:
            logger.error(f"Failed to write to Amazon template Excel: {e}")
            raise


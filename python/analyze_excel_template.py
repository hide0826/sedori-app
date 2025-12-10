#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path
import json

file_path = Path(r"D:\HIRIO\repo\sedori-app.github\python\ListingLoader.xlsm")

result = {
    "file_path": str(file_path),
    "file_exists": file_path.exists(),
    "error": None,
    "sheets": [],
    "template_sheet": None
}

try:
    if not file_path.exists():
        result["error"] = "File not found"
    else:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        result["sheets"] = wb.sheetnames
        
        if "テンプレート" in wb.sheetnames:
            ws = wb["テンプレート"]
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(1, col).value
                if val:
                    headers.append({"column": col, "header": str(val)})
            
            # サンプルデータ（2-5行目）
            sample_data = []
            for row in range(2, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(ws.max_column + 1, 20)):  # 最初の20列まで
                    val = ws.cell(row, col).value
                    row_data.append(str(val) if val is not None else "")
                sample_data.append(row_data)
            
            result["template_sheet"] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": headers,
                "sample_data": sample_data
            }
        else:
            result["error"] = "Template sheet not found"
    
    # JSON形式で保存
    output_file = Path(__file__).parent / "excel_template_result.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    # テキスト形式でも保存（読みやすいように）
    output_txt = Path(__file__).parent / "excel_template_result.txt"
    with open(output_txt, "w", encoding="utf-8") as f:
        f.write("=== Excel Template Analysis ===\n\n")
        f.write(f"File: {result['file_path']}\n")
        f.write(f"File exists: {result['file_exists']}\n\n")
        
        if result["error"]:
            f.write(f"Error: {result['error']}\n\n")
        
        f.write(f"Available sheets: {', '.join(result['sheets'])}\n\n")
        
        if result["template_sheet"]:
            ts = result["template_sheet"]
            f.write("=== Template Sheet ===\n")
            f.write(f"Max row: {ts['max_row']}\n")
            f.write(f"Max column: {ts['max_column']}\n\n")
            f.write("Headers:\n")
            for h in ts["headers"]:
                f.write(f"  Column {h['column']}: {h['header']}\n")
            f.write("\nSample data (rows 2-5):\n")
            for idx, row_data in enumerate(ts["sample_data"], start=2):
                f.write(f"  Row {idx}: {row_data}\n")
    
    print("Analysis complete. Results saved to:")
    print(f"  - {output_file}")
    print(f"  - {output_txt}")
    
except Exception as e:
    result["error"] = str(e)
    import traceback
    result["traceback"] = traceback.format_exc()
    
    output_file = Path(__file__).parent / "excel_template_result.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"Error: {e}")
    traceback.print_exc()





#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AmazonテンプレートExcelファイルの「テンプレート」シートを読み込むスクリプト
"""
import openpyxl
from pathlib import Path
import sys

output_lines = []

def log(msg):
    output_lines.append(msg)
    print(msg)
    sys.stdout.flush()

try:
    # Excelファイルを読み込む
    file_path = Path(r"D:\HIRIO\repo\sedori-app.github\python\ListingLoader.xlsm")
    log(f"ファイルを読み込み中: {file_path}")
    
    if not file_path.exists():
        log(f"エラー: ファイルが見つかりません: {file_path}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(file_path, keep_vba=True)  # keep_vba=Trueでマクロ有効ファイルを読み込み
    log("ファイルの読み込みが完了しました。")
    
    # シート名を確認
    log("\n利用可能なシート名:")
    for sheet_name in wb.sheetnames:
        log(f"  - {sheet_name}")
    
    # 「テンプレート」シートを取得
    if "テンプレート" in wb.sheetnames:
        ws = wb["テンプレート"]
        log(f"\n「テンプレート」シートを読み込みました。")
        log(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
        
        # ヘッダー行（1行目）を確認
        log("\n【ヘッダー行（1行目）】")
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                headers.append((col_idx, cell_value))
                log(f"  列{col_idx}: {cell_value}")
        
        # データ行のサンプル（2-5行目）を確認
        log("\n【データ行のサンプル（2-5行目）】")
        for row_idx in range(2, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            log(f"  行{row_idx}: {row_data[:10]}...")  # 最初の10列だけ表示
        
        # 結果をファイルに保存
        output_file = Path(__file__).parent / "excel_template_analysis.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))
        log(f"\n結果をファイルに保存しました: {output_file}")
        
    else:
        log("\n「テンプレート」という名前のシートが見つかりませんでした。")
        log("利用可能なシート名を確認してください。")
        
except Exception as e:
    log(f"エラーが発生しました: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)









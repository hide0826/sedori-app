import sys
print("Python version:", sys.version)
print("Starting...")

try:
    import openpyxl
    print("openpyxl imported successfully")
    print("openpyxl version:", openpyxl.__version__)
except ImportError as e:
    print("openpyxl not installed:", e)
    sys.exit(1)

from pathlib import Path
file_path = Path(r"D:\HIRIO\repo\sedori-app.github\python\ListingLoader.xlsm")
print(f"File exists: {file_path.exists()}")
print(f"File path: {file_path}")

if file_path.exists():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    print("Workbook loaded!")
    print("Sheet names:", wb.sheetnames)
    
    if "テンプレート" in wb.sheetnames:
        ws = wb["テンプレート"]
        print(f"Template sheet found! Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # ヘッダーを取得
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):  # 最大50列まで
            val = ws.cell(1, col).value
            if val:
                headers.append((col, val))
                print(f"Col {col}: {val}")
        
        # 結果をファイルに保存
        with open("excel_result.txt", "w", encoding="utf-8") as f:
            f.write(f"Sheet names: {wb.sheetnames}\n")
            f.write(f"Template sheet - Max row: {ws.max_row}, Max col: {ws.max_column}\n")
            f.write("\nHeaders:\n")
            for col, val in headers:
                f.write(f"Col {col}: {val}\n")
        print("Results saved to excel_result.txt")
    else:
        print("Template sheet not found")





#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelインポート機能

仕入先マスタシート（Excel）から店舗データを読み込む
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl


class ExcelImporter:
    """Excelインポート機能クラス"""
    
    def __init__(self):
        """初期化"""
        pass
    
    def read_store_master_sheet(self, excel_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """仕入先マスタシートを読み込む"""
        try:
            # シート名が指定されていない場合は最初のシートを使用
            if sheet_name is None:
                # Excelファイルからシート名を取得
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                sheet_names = wb.sheetnames
                
                # 「仕入先マスタ」という名前のシートを探す
                if "仕入先マスタ" in sheet_names:
                    sheet_name = "仕入先マスタ"
                else:
                    # 見つからない場合は最初のシートを使用
                    sheet_name = sheet_names[0] if sheet_names else None
                    if not sheet_name:
                        raise ValueError("Excelファイルにシートが見つかりません")
            
            # Excelファイルを読み込む
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                dtype=str,  # すべて文字列として読み込み
                keep_default_na=False  # NaNを空文字列として扱う
            )
            
            # 列名を正規化（空白除去、前後空白削除）
            df.columns = df.columns.str.strip()
            
            # データを辞書のリストに変換
            stores = []
            for _, row in df.iterrows():
                store = self._parse_store_row(row)
                if store:  # 空行はスキップ
                    stores.append(store)
            
            return stores
            
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みに失敗しました: {str(e)}")
    
    def _parse_store_row(self, row: pd.Series) -> Optional[Dict[str, Any]]:
        """行データを店舗辞書に変換"""
        # 列名マッピング（複数の可能性のある列名に対応）
        column_mapping = {
            '所属ルート名': ['所属ルート名', 'ルート名', 'route_name'],
            'ルートコード': ['ルートコード', 'route_code', 'コード'],
            '仕入れ先コード': ['仕入れ先コード', '仕入先コード', 'supplier_code', '店舗コード'],
            '店舗名': ['店舗名', 'store_name', '名前']
        }
        
        store = {}
        
        # マッピングに基づいてデータを取得
        for key, possible_names in column_mapping.items():
            value = None
            for name in possible_names:
                if name in row.index:
                    value = str(row[name]).strip()
                    if value and value != 'nan':
                        break
            
            # キー名を変換（スネークケース）
            if key == '所属ルート名':
                store['affiliated_route_name'] = value or ''
            elif key == 'ルートコード':
                store['route_code'] = value or ''
            elif key == '仕入れ先コード':
                store['supplier_code'] = value or ''
            elif key == '店舗名':
                store['store_name'] = value or ''
        
        # 店舗名が必須（空の場合はスキップ）
        if not store.get('store_name'):
            return None
        
        # 仕入れ先コードがない場合は警告（必須ではない）
        if not store.get('supplier_code'):
            # 仕入れ先コードがない場合は警告のみ（スキップしない）
            pass
        
        # カスタムフィールドは空で初期化
        store['custom_fields'] = {}
        
        return store
    
    def get_excel_sheet_names(self, excel_path: str) -> List[str]:
        """Excelファイル内のシート名一覧を取得"""
        try:
            wb = openpyxl.load_workbook(excel_path)
            return wb.sheetnames
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みに失敗しました: {str(e)}")
    
    def validate_store_data(self, stores: List[Dict[str, Any]]) -> Dict[str, Any]:
        """店舗データの検証"""
        errors = []
        warnings = []
        valid_count = 0
        
        for i, store in enumerate(stores, start=2):  # Excelの行番号（1行目はヘッダー）
            # 必須項目チェック
            if not store.get('store_name'):
                errors.append(f"行{i}: 店舗名が空です")
                continue
            
            # 仕入れ先コードの重複チェック（同じリスト内）
            supplier_code = store.get('supplier_code')
            if supplier_code:
                duplicate_count = sum(
                    1 for s in stores if s.get('supplier_code') == supplier_code
                )
                if duplicate_count > 1:
                    warnings.append(f"行{i}: 仕入れ先コード '{supplier_code}' が重複しています")
            
            valid_count += 1
        
        return {
            'valid_count': valid_count,
            'error_count': len(errors),
            'warning_count': len(warnings),
            'errors': errors,
            'warnings': warnings
        }


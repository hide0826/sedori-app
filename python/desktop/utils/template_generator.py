#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ルートサマリーテンプレート生成ユーティリティ

ルートサマリー入力用のCSV/Excelテンプレートを生成
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class TemplateGenerator:
    """ルートサマリーテンプレート生成クラス"""
    
    @staticmethod
    def generate_csv_template(output_path: str, route_code: Optional[str] = None, store_codes: Optional[List[str]] = None, route_date: Optional[datetime.date] = None) -> bool:
        """
        CSV形式のルートサマリーテンプレートを生成
        
        Args:
            output_path: 出力ファイルパス
            route_code: ルートコード（任意）
            store_codes: 店舗コードリスト（任意）
            route_date: ルート日付（任意、指定されない場合は今日の日付を使用）
        
        Returns:
            生成成功時True
        """
        try:
            # ルート日付を取得（指定されていない場合は今日の日付）
            route_date_str = route_date.strftime('%Y-%m-%d') if route_date else datetime.now().strftime('%Y-%m-%d')
            
            # ルート情報セクションのデータ
            route_data = {
                '項目': [
                    'ルート日付',
                    'ルートコード',
                    '出発時間',
                    '帰宅時間',
                    '往路高速代',
                    '復路高速代',
                    '駐車場代',
                    '食費',
                    'その他経費',
                    '備考（天候等）'
                ],
                '値': [
                    route_date_str,
                    route_code or '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ]
            }
            
            # 店舗訪問詳細セクションのヘッダー
            visit_headers = [
                '訪問順序',
                '店舗コード',
                '店舗IN時間',
                '店舗OUT時間',
                '前店舗からの移動時間（分）',
                '前店舗からの距離（km）',
                '店舗毎想定粗利',
                '店舗毎仕入れ点数',
                '仕入れ成功',
                '空振り理由',
                '店舗評価（1-5）',
                '店舗メモ',
                '次回訪問推奨度',
                '在庫状況',
                '競合いたか',
                'トラブルあったか',
                'トラブル詳細'
            ]
            
            # 店舗訪問詳細の初期データ
            visit_rows = []
            if store_codes:
                for i, store_code in enumerate(store_codes, 1):
                    visit_rows.append([
                        i,  # 訪問順序
                        store_code,
                        '',  # 店舗IN時間
                        '',  # 店舗OUT時間
                        '',  # 移動時間
                        '',  # 距離
                        '',  # 想定粗利
                        '',  # 仕入れ点数
                        'YES',  # 仕入れ成功（デフォルト）
                        '',  # 空振り理由
                        '',  # 店舗評価
                        '',  # 店舗メモ
                        '',  # 次回訪問推奨度
                        '',  # 在庫状況
                        'NO',  # 競合
                        'NO',  # トラブル
                        ''  # トラブル詳細
                    ])
            else:
                # 店舗コードが指定されていない場合は空の行を5行追加
                for i in range(1, 6):
                    visit_rows.append([
                        i,
                        '',
                        '',
                        '',
                        '',
                        '',
                        '',
                        '',
                        'YES',
                        '',
                        '',
                        '',
                        '',
                        '',
                        'NO',
                        'NO',
                        ''
                    ])
            
            # CSV形式で出力
            output_lines = []
            
            # ルート情報セクション
            output_lines.append('# === ルート情報 ===')
            output_lines.append(','.join(route_data['項目']))
            output_lines.append(','.join(str(v) for v in route_data['値']))
            output_lines.append('')
            
            # 店舗訪問詳細セクション
            output_lines.append('# === 店舗訪問詳細 ===')
            output_lines.append(','.join(visit_headers))
            for row in visit_rows:
                output_lines.append(','.join(str(v) for v in row))
            
            # ファイルに書き込み（UTF-8 BOM付き）
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                f.write('\n'.join(output_lines))
            
            return True
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"CSVテンプレート生成エラー: {e}")
            print(f"エラー詳細:\n{error_detail}")
            return False
    
    @staticmethod
    def generate_excel_template(output_path: str, route_code: Optional[str] = None, store_codes: Optional[List[str]] = None, stores_data: Optional[List[Dict[str, Any]]] = None, route_date: Optional[datetime.date] = None) -> bool:
        """
        Excel形式のルートサマリーテンプレートを生成
        
        Args:
            output_path: 出力ファイルパス
            route_code: ルートコード（任意）
            store_codes: 店舗コードリスト（任意）
            stores_data: 店舗データリスト（店舗名等を含む）
            route_date: ルート日付（任意、指定されない場合は今日の日付を使用）
        
        Returns:
            生成成功時True
        """
        try:
            print(f"テンプレート生成開始: {output_path}")
            print(f"ルートコード: {route_code}, 店舗数: {len(stores_data) if stores_data else len(store_codes) if store_codes else 0}")
            
            # ワークブック作成
            wb = openpyxl.Workbook()
            
            # デフォルトシートを削除
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # スタイル定義
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # === 店舗訪問詳細（単一シート） ===
            visit_sheet = wb.create_sheet('店舗訪問詳細', 0)
            
            # 上部情報（日付・ルート名）
            route_date_cell = visit_sheet.cell(row=1, column=1)
            route_date_cell.value = '日付'
            route_date_cell.fill = header_fill
            route_date_cell.font = header_font
            route_date_cell.border = border
            route_date_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            date_value_cell = visit_sheet.cell(row=1, column=2)
            # ルート日付が指定されている場合はそれを使用、そうでない場合は今日の日付を使用
            date_value_cell.value = route_date if route_date else datetime.now().date()
            date_value_cell.border = border
            date_value_cell.number_format = 'yyyy/mm/dd'  # 日付形式
            
            route_name_cell = visit_sheet.cell(row=2, column=1)
            route_name_cell.value = 'ルート'
            route_name_cell.fill = header_fill
            route_name_cell.font = header_font
            route_name_cell.border = border
            route_name_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            route_value_cell = visit_sheet.cell(row=2, column=2)
            route_value_cell.value = route_code or ''
            route_value_cell.border = border
            
            # ヘッダー（3行目） - 日付列を削除
            visit_headers = [
                '店舗コード',
                '店舗名',
                '到着時刻',
                '出発時刻',
                '滞在時間',
                '備考'
            ]
            
            for col_idx, header in enumerate(visit_headers, start=1):
                cell = visit_sheet.cell(row=3, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # データ行（4行目以降）
            visit_data_rows = []
            if stores_data:
                # 店舗データを使用してテンプレートを生成
                # 店舗コードは store_code を優先し、なければ互換性のため supplier_code を使用
                for i, store in enumerate(stores_data, 1):
                    # 店舗マスタの備考欄からメモを取得
                    notes = store.get('notes', '')
                    store_code = store.get('store_code') or store.get('supplier_code') or ''
                    visit_data_rows.append([
                        store_code,                    # 店舗コード
                        store.get('store_name', ''),   # 店舗名
                        '',                            # 到着時刻
                        '',                            # 出発時刻
                        '0',                           # 滞在時間（数式で自動計算される）
                        notes                          # 備考（店舗マスタから取得）
                    ])
            elif store_codes:
                # 店舗コードのみの場合
                for i, store_code in enumerate(store_codes, 1):
                    visit_data_rows.append([
                        store_code,
                        '',  # 店舗名
                        '',  # 到着時刻
                        '',  # 出発時刻
                        '0',  # 滞在時間
                        ''   # 備考
                    ])
            else:
                # 空の行を追加
                for i in range(1, 6):
                    visit_data_rows.append([
                        '',
                        '',
                        '',
                        '',
                        '0',
                        ''
                    ])
            
            for row_idx, row_data in enumerate(visit_data_rows, start=4):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = visit_sheet.cell(row=row_idx, column=col_idx)
                    if col_idx == 5:  # 滞在時間列（E列）
                        # 到着時刻（C列）と出発時刻（D列）から分数を計算する数式
                        # 時刻が入力されている場合のみ計算
                        arrival_col = get_column_letter(3)  # C列
                        departure_col = get_column_letter(4)  # D列
                        formula = f'=IF(AND({arrival_col}{row_idx}<>"",{departure_col}{row_idx}<>""),ROUND(({departure_col}{row_idx}-{arrival_col}{row_idx})*24*60,0),0)'
                        cell.value = formula
                        cell.number_format = '0'  # 整数表示
                    elif col_idx == 3 or col_idx == 4:  # 到着時刻・出発時刻列
                        cell.value = value
                        cell.number_format = 'hh:mm'  # 時刻形式
                    else:
                        cell.value = value
                    cell.border = border
            
            # 店舗データの直後に追加情報を配置（店舗データの最後の行の次から）
            # 店舗データは4行目から始まるので、最後の行番号は 3 + len(visit_data_rows)
            # その次の行（4 + len(visit_data_rows)）から追加情報を配置
            info_start_row = 4 + len(visit_data_rows)
            info_labels = ['出発時刻', '帰宅時刻', '往路高速代', '復路高速代']
            for i, label in enumerate(info_labels):
                row_num = info_start_row + i
                label_cell = visit_sheet.cell(row=row_num, column=1)
                label_cell.value = label
                label_cell.fill = header_fill
                label_cell.font = header_font
                label_cell.border = border
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 値セル（B列に配置）
                value_cell = visit_sheet.cell(row=row_num, column=2)
                value_cell.value = ''
                value_cell.border = border
                
                # 時刻関連のラベルは時刻形式を設定
                if label in ['出発時刻', '帰宅時刻']:
                    value_cell.number_format = 'hh:mm'
            
            # 列幅調整（6列構成）
            column_widths = [15, 40, 18, 18, 12, 30]
            for col_idx, width in enumerate(column_widths, start=1):
                visit_sheet.column_dimensions[get_column_letter(col_idx)].width = width
            
            # 3行目を固定
            visit_sheet.freeze_panes = 'A4'
            
            # 保存
            wb.save(output_path)
            print(f"テンプレート生成完了: {output_path}")
            print(f"  シート数: {len(wb.sheetnames)}")
            print(f"  シート名: {wb.sheetnames}")
            return True
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"Excelテンプレート生成エラー: {e}")
            print(f"エラー詳細:\n{error_detail}")
            return False


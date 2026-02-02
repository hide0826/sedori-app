#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ルートサマリー入力ウィジェット

ルートサマリーの入力・編集機能
- テンプレートファイルの読み込み
- ルート情報の入力・編集
- 店舗別情報の入力・編集
- 照合処理・店舗コード自動挿入
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QMessageBox, QFileDialog, QDateTimeEdit, QLineEdit,
    QTextEdit, QDoubleSpinBox, QSpinBox, QCheckBox, QComboBox,
    QDialog, QFormLayout, QDialogButtonBox, QTabWidget, QStyledItemDelegate, QStyle, QInputDialog
)
from ui.star_rating_widget import StarRatingWidget
from PySide6.QtCore import Qt, QDateTime, QTime, Signal, QSettings
from PySide6.QtGui import QColor, QShortcut, QKeySequence
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import sys
import os
from datetime import time as dt_time
import openpyxl
import logging
from functools import partial

# プロジェクトルートをパスに追加
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from database.route_db import RouteDatabase
from database.store_db import StoreDatabase
from database.route_visit_db import RouteVisitDatabase

# サービス・ユーティリティのインポート（相対パス）
import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    from services.route_matching_service import RouteMatchingService
except ImportError:
    RouteMatchingService = None  # オプション機能として扱う

try:
    from services.calculation_service import CalculationService
except ImportError:
    CalculationService = None

try:
    from utils.template_generator import TemplateGenerator
except ImportError:
    TemplateGenerator = None


class RouteSummaryWidget(QWidget):
    """ルートサマリー入力ウィジェット"""
    
    data_saved = Signal(int)  # データ保存完了
    
    def __init__(self, api_client=None, inventory_widget=None):
        super().__init__()
        self.route_db = RouteDatabase()
        self.store_db = StoreDatabase()
        self.route_visit_db = RouteVisitDatabase()
        self.matching_service = RouteMatchingService() if RouteMatchingService else None
        self.calc_service = CalculationService() if CalculationService else None
        self.api_client = api_client
        self.inventory_widget = inventory_widget
        
        self.current_route_id = None
        self.route_data = {}
        self.store_visits = []
        self.latest_summary_metrics = {}
        self.settings = QSettings("HIRIO", "SedoriDesktopApp")
        self.last_loaded_template_path: str = ""
        # テンプレート生成のデフォルトフォルダ
        stored_template_dir = self.settings.value("route_template/default_save_dir", "")
        self.template_save_default_dir = str(stored_template_dir) if stored_template_dir else ""
        
        # Undo/Redoスタック
        self.undo_stack = []
        self.redo_stack = []
        self.max_undo_history = 50  # 最大履歴数
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIの設定"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # 上部：操作ボタン
        self.setup_action_buttons(layout)
        
        # 中央：タブ（ルート情報・店舗訪問詳細）
        self.setup_tabs(layout)
        
        # 下部：計算結果表示
        self.setup_calculation_results(layout)
    
    def setup_action_buttons(self, parent_layout):
        """操作ボタンの設定"""
        button_group = QGroupBox("操作")
        button_layout = QHBoxLayout(button_group)
        
        # テンプレート生成ボタン
        template_btn = QPushButton("テンプレート生成")
        template_btn.clicked.connect(self.generate_template)
        template_btn.setStyleSheet("background-color: #28a745; color: white;")
        button_layout.addWidget(template_btn)
        
        # 選択ルート読み込みボタン
        auto_add_btn = QPushButton("選択ルート読み込み")
        auto_add_btn.clicked.connect(self.auto_add_stores)
        auto_add_btn.setStyleSheet("background-color: #17a2b8; color: white;")
        button_layout.addWidget(auto_add_btn)
        
        button_layout.addStretch()
        
        # テンプレート生成フォルダ指定（右側に配置）
        template_save_dir_label = QLabel("テンプレート保存フォルダ:")
        template_save_dir_label.setStyleSheet("color: #cccccc;")
        button_layout.addWidget(template_save_dir_label)
        
        self.template_save_dir_edit = QLineEdit()
        self.template_save_dir_edit.setPlaceholderText("未設定")
        self.template_save_dir_edit.setFixedWidth(260)
        self.template_save_dir_edit.setText(self.template_save_default_dir)
        self.template_save_dir_edit.editingFinished.connect(self.on_template_save_dir_edit_finished)
        button_layout.addWidget(self.template_save_dir_edit)
        
        browse_save_dir_btn = QPushButton("参照…")
        browse_save_dir_btn.clicked.connect(self.browse_template_save_dir)
        button_layout.addWidget(browse_save_dir_btn)
        
        parent_layout.addWidget(button_group)
    
    def setup_tabs(self, parent_layout):
        """タブの設定（統合版）"""
        # 単一のウィジェットに統合
        unified_widget = self.create_unified_widget()
        parent_layout.addWidget(unified_widget)
    
    def create_unified_widget(self) -> QWidget:
        """ルート情報と店舗訪問詳細を統合したウィジェット"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(10)
        
        # 上部：ルート情報セクション（サイズを小さく設定）
        route_group = QGroupBox("ルート情報")
        route_group.setMaximumHeight(300)  # 最大高さを制限
        route_layout = QFormLayout(route_group)
        route_layout.setSpacing(8)  # スペーシングを詰める
        
        # ルート日付
        self.route_date_edit = QDateTimeEdit()
        self.route_date_edit.setCalendarPopup(True)
        self.route_date_edit.setDateTime(QDateTime.currentDateTime())
        self.route_date_edit.setDisplayFormat("yyyy-MM-dd")
        route_layout.addRow("ルート日付:", self.route_date_edit)
        
        # ルートコード（プルダウン）
        self.route_code_combo = QComboBox()
        self.route_code_combo.setEditable(True)
        self.route_code_combo.currentTextChanged.connect(self.on_route_code_changed)
        route_layout.addRow("ルートコード:", self.route_code_combo)
        self.update_route_codes()
        # デフォルトは空白（update_route_codes()の後に設定）
        self.route_code_combo.setCurrentText("")
        
        layout.addWidget(route_group, 0)  # stretch=0で最小サイズに
        
        # 下部：店舗訪問詳細セクション（残りのスペースを全て使用）
        visits_group = QGroupBox("店舗訪問詳細")
        visits_layout = QVBoxLayout(visits_group)
        
        # テーブル
        self.store_visits_table = QTableWidget()
        self.store_visits_table.setAlternatingRowColors(True)
        # 行ドラッグ＆ドロップに必要な行単位選択＋単一選択
        self.store_visits_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.store_visits_table.setSelectionMode(QTableWidget.SingleSelection)
        self.store_visits_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed)
        # テキストの省略（...）を無効化
        self.store_visits_table.setTextElideMode(Qt.ElideNone)
        # テキストを折り返して全文表示
        self.store_visits_table.setWordWrap(True)
        
        # ドラッグ＆ドロップを有効化（行間に挿入する設定）
        self.store_visits_table.setDragEnabled(True)
        self.store_visits_table.setAcceptDrops(True)
        self.store_visits_table.setDropIndicatorShown(True)
        self.store_visits_table.setDragDropMode(QTableWidget.InternalMove)
        self.store_visits_table.setDragDropOverwriteMode(False)
        # ドラッグ中の自動スクロールが列ズレの原因になるため無効化
        try:
            from PySide6.QtWidgets import QAbstractItemView
            self.store_visits_table.setAutoScroll(False)
            self.store_visits_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        except Exception:
            pass
        self.store_visits_table.setDefaultDropAction(Qt.MoveAction)
        
        headers = [
            "訪問順序", "店舗コード", "店舗名", "店舗IN時間", "店舗OUT時間",
            "店舗滞在時間", "移動時間（分）", "想定粗利", "仕入れ点数",
            "店舗評価", "備考"
        ]
        self.store_visits_table.setColumnCount(len(headers))
        self.store_visits_table.setHorizontalHeaderLabels(headers)
        
        header = self.store_visits_table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.Interactive)
        # 列のドラッグ移動は不可（行ドラッグ時の横スライドで列が入れ替わらないよう固定）
        try:
            header.setSectionsMovable(False)
        except Exception:
            pass
        try:
            self.store_visits_table.verticalHeader().setSectionsMovable(False)
        except Exception:
            pass
        
        # 店舗名列をコンテンツに合わせて自動調整
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 店舗名列（インデックス2）
        # 備考カラム（10列目）はStretchモードで残りのスペースを使用
        header.setSectionResizeMode(10, QHeaderView.Stretch)
        
        # デフォルトの行高を調整（星評価が綺麗に収まるように）
        self.store_visits_table.verticalHeader().setDefaultSectionSize(24)
        
        # 訪問順序の変更を監視
        self.store_visits_table.model().rowsMoved.connect(self.on_rows_moved)
        
        # データ変更を監視してUndoスタックに保存
        self.store_visits_table.itemChanged.connect(self.on_table_item_changed)
        
        # ショートカットキーの設定
        self.setup_shortcuts()
        
        # メモ列のフォーカス・選択の強調を最小化（実質非表示）
        class MinimalFocusDelegate(QStyledItemDelegate):
            def paint(self, painter, option, index):
                if option.state & QStyle.State_HasFocus:
                    option.state &= ~QStyle.State_HasFocus
                if option.state & QStyle.State_Selected:
                    option.state &= ~QStyle.State_Selected
                super().paint(painter, option, index)
            
            def createEditor(self, parent, option, index):
                # メモ入力用のエディタ（枠線なし・背景そのまま）
                from PySide6.QtWidgets import QLineEdit
                editor = QLineEdit(parent)
                editor.setFrame(False)
                # 入力中の視認性を上げるため、セル内に小さな編集ボックス風の装飾を適用
                # 背景は不透明にして下地テキストが透けないようにする（ゴースト文字対策）
                editor.setStyleSheet(
                    "QLineEdit{"
                    "background-color: #2b2b2b;"
                    "color: #ffffff;"
                    "border: 1px solid #5aa2ff;"
                    "border-radius: 3px;"
                    "padding: 2px 6px;"
                    "}"
                )
                try:
                    editor.setAutoFillBackground(True)
                except Exception:
                    pass
                try:
                    editor.setTextMargins(4, 0, 4, 0)
                except Exception:
                    pass
                return editor

        self.store_visits_table.setItemDelegateForColumn(10, MinimalFocusDelegate(self.store_visits_table))

        # 選択時の可視性を上げつつ、強すぎないハイライトに調整
        self.store_visits_table.setStyleSheet(
            "QTableView::item:focus{outline: none;}"
            "QTableView::item:selected{background-color: rgba(90,162,255,0.18); color: #ffffff; border: none;}"
            "QTableView::item{border: none;}"
        )


        # ドラッグ＆ドロップのため選択は有効（単一行）
        visits_layout.addWidget(self.store_visits_table)
        
        # 行追加・削除・全クリア・Undo/Redoボタン
        button_layout = QHBoxLayout()
        # 店舗マスタから追加ボタン
        add_from_master_btn = QPushButton("店舗追加（店舗マスタから）")
        add_from_master_btn.clicked.connect(self.add_store_from_master)
        button_layout.addWidget(add_from_master_btn)
        add_row_btn = QPushButton("行追加")
        add_row_btn.clicked.connect(self.add_store_visit_row)
        button_layout.addWidget(add_row_btn)
        
        delete_row_btn = QPushButton("行削除")
        delete_row_btn.clicked.connect(self.delete_store_visit_row)
        button_layout.addWidget(delete_row_btn)
        
        clear_all_btn = QPushButton("全行クリア")
        clear_all_btn.clicked.connect(self.clear_all_rows)
        clear_all_btn.setStyleSheet("background-color: #dc3545; color: white;")
        button_layout.addWidget(clear_all_btn)
        
        # 一括入替ボタン（訪問順序に基づいて行を並び替え）
        reorder_btn = QPushButton("一括入替")
        reorder_btn.clicked.connect(self.reorder_by_visit_order)
        reorder_btn.setToolTip("訪問順序列に入力した数字に基づいて行を並び替えます")
        reorder_btn.setStyleSheet("background-color: #17a2b8; color: white;")
        button_layout.addWidget(reorder_btn)
        
        # 訪問順序保存ボタン
        save_order_btn = QPushButton("訪問順序保存")
        save_order_btn.clicked.connect(self.save_visit_order_to_db)
        save_order_btn.setToolTip("現在の訪問順序をデータベースに保存します。次回同じルートを呼び出した時に保存された順序で表示されます。")
        save_order_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        button_layout.addWidget(save_order_btn)
        
        button_layout.addStretch()
        
        # Undo/Redoボタン
        undo_btn = QPushButton("元に戻す (Ctrl+Z)")
        undo_btn.clicked.connect(self.undo_action)
        undo_btn.setStyleSheet("background-color: #6c757d; color: white;")
        button_layout.addWidget(undo_btn)
        
        redo_btn = QPushButton("やり直す (Ctrl+Y)")
        redo_btn.clicked.connect(self.redo_action)
        redo_btn.setStyleSheet("background-color: #6c757d; color: white;")
        button_layout.addWidget(redo_btn)
        
        visits_layout.addLayout(button_layout)
        layout.addWidget(visits_group, 1)  # stretch=1で残りのスペースを全て使用
        
        # 初期スナップショットを保存（Undoの基点）
        try:
            self.save_table_state()
        except Exception:
            pass
        return widget
    
    def showEvent(self, event):
        """ウィジェットが表示されたときにルートコード一覧を更新"""
        super().showEvent(event)
        # ルートコード一覧を最新の状態に更新
        if hasattr(self, 'route_code_combo') and self.route_code_combo:
            self.update_route_codes()
    
    def add_store_from_master(self):
        """店舗マスタ一覧から選択して店舗を追加"""
        try:
            dialog = StoreSelectDialog(self.store_db, self)
            if dialog.exec() == QDialog.Accepted:
                selected = dialog.get_selected_stores()
                if not selected:
                    return
                # 変更前の状態保存
                self.save_table_state()
                base_rows = self.store_visits_table.rowCount()
                for i, store in enumerate(selected):
                    row = base_rows + i
                    self.store_visits_table.insertRow(row)
                    order_item = QTableWidgetItem(str(row + 1))
                    order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                    self.store_visits_table.setItem(row, 0, order_item)
                    # 店舗コード列には店舗マスタの「店舗コード(store_code)」を優先して使用し、
                    # 空の場合は仕入れ先コード（supplier_code）をフォールバックとして使用する
                    store_code = (store.get('store_code') or '').strip()
                    visit_code = store_code or (store.get('supplier_code') or '').strip()
                    self.store_visits_table.setItem(row, 1, QTableWidgetItem(visit_code))
                    self.store_visits_table.setItem(row, 2, QTableWidgetItem(store.get('store_name', '')))
                    # 星評価初期化
                    star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
                    star_widget.rating_changed.connect(lambda rating, r=row: self.on_star_rating_changed(r, rating))
                    self.store_visits_table.setCellWidget(row, 9, star_widget)
                self.update_visit_order()
                self.save_store_order()
                self.recalc_travel_times()
                self.save_table_state()
                # 計算結果更新（存在しない環境でも落ちないように）
                getattr(self, 'update_calculation_results', lambda: None)()
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"店舗追加中にエラーが発生しました:\n{str(e)}")
    
    def create_route_info_widget(self) -> QWidget:
        """ルート情報入力ウィジェットの作成"""
        widget = QWidget()
        layout = QFormLayout(widget)
        layout.setSpacing(10)
        
        # ルート日付
        self.route_date_edit = QDateTimeEdit()
        self.route_date_edit.setCalendarPopup(True)
        self.route_date_edit.setDateTime(QDateTime.currentDateTime())
        self.route_date_edit.setDisplayFormat("yyyy-MM-dd")
        layout.addRow("ルート日付:", self.route_date_edit)
        
        # ルートコード（プルダウン）
        self.route_code_combo = QComboBox()
        self.route_code_combo.setEditable(True)  # 手動入力も可能
        self.route_code_combo.currentTextChanged.connect(self.on_route_code_changed)
        layout.addRow("ルートコード:", self.route_code_combo)
        
        # ルートコード一覧を更新
        self.update_route_codes()
        # デフォルトは空白（update_route_codes()の後に設定）
        self.route_code_combo.setCurrentText("")
        
        return widget
    
    def create_store_visits_widget(self) -> QWidget:
        """店舗訪問詳細テーブルウィジェットの作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # テーブル
        self.store_visits_table = QTableWidget()
        self.store_visits_table.setAlternatingRowColors(True)
        self.store_visits_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.store_visits_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed)
        # テキストの省略（...）を無効化
        self.store_visits_table.setTextElideMode(Qt.ElideNone)
        # テキストを折り返して全文表示
        self.store_visits_table.setWordWrap(True)
        
        # ドラッグ＆ドロップを有効化（行間に挿入する設定）
        self.store_visits_table.setDragEnabled(True)
        self.store_visits_table.setAcceptDrops(True)
        self.store_visits_table.setDropIndicatorShown(True)
        self.store_visits_table.setDragDropMode(QTableWidget.InternalMove)
        self.store_visits_table.setDragDropOverwriteMode(False)
        self.store_visits_table.setDefaultDropAction(Qt.MoveAction)
        
        headers = [
            "訪問順序", "店舗コード", "店舗名", "店舗IN時間", "店舗OUT時間",
            "店舗滞在時間", "移動時間（分）", "想定粗利", "仕入れ点数",
            "店舗評価", "備考"
        ]
        self.store_visits_table.setColumnCount(len(headers))
        self.store_visits_table.setHorizontalHeaderLabels(headers)
        
        header = self.store_visits_table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.Interactive)
        
        # 店舗名列をコンテンツに合わせて自動調整
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 店舗名列（インデックス2）
        # 備考カラム（10列目）はStretchモードで残りのスペースを使用
        header.setSectionResizeMode(10, QHeaderView.Stretch)
        
        # デフォルトの行高を調整（星評価が綺麗に収まるように）
        self.store_visits_table.verticalHeader().setDefaultSectionSize(24)
        
        # 訪問順序の変更を監視
        self.store_visits_table.model().rowsMoved.connect(self.on_rows_moved)
        
        # データ変更を監視してUndoスタックに保存
        self.store_visits_table.itemChanged.connect(self.on_table_item_changed)
        
        # ショートカットキーの設定
        self.setup_shortcuts()
        
        layout.addWidget(self.store_visits_table)
        
        # 行追加・削除・全クリア・Undo/Redoボタン
        button_layout = QHBoxLayout()
        add_row_btn = QPushButton("行追加")
        add_row_btn.clicked.connect(lambda: getattr(self, 'add_store_visit_row', lambda: None)())
        button_layout.addWidget(add_row_btn)
        
        delete_row_btn = QPushButton("行削除")
        delete_row_btn.clicked.connect(lambda: getattr(self, 'delete_store_visit_row', lambda: None)())
        button_layout.addWidget(delete_row_btn)
        
        clear_all_btn = QPushButton("全行クリア")
        clear_all_btn.clicked.connect(lambda: getattr(self, 'clear_all_rows', lambda: None)())
        clear_all_btn.setStyleSheet("background-color: #dc3545; color: white;")
        button_layout.addWidget(clear_all_btn)
        
        button_layout.addStretch()
        
        # Undo/Redoボタン
        undo_btn = QPushButton("元に戻す (Ctrl+Z)")
        undo_btn.clicked.connect(self.undo_action)
        undo_btn.setStyleSheet("background-color: #6c757d; color: white;")
        button_layout.addWidget(undo_btn)
        
        redo_btn = QPushButton("やり直す (Ctrl+Y)")
        redo_btn.clicked.connect(self.redo_action)
        redo_btn.setStyleSheet("background-color: #6c757d; color: white;")
        button_layout.addWidget(redo_btn)
        
        layout.addLayout(button_layout)
        
        return widget
    
    def setup_calculation_results(self, parent_layout):
        """計算結果表示の設定（非表示）"""
        result_group = QGroupBox("計算結果")
        result_layout = QVBoxLayout(result_group)
        
        self.calculation_label = QLabel("計算結果: データ未入力")
        result_layout.addWidget(self.calculation_label)
        
        # 計算結果エリアを非表示にする
        result_group.setVisible(False)
        
        parent_layout.addWidget(result_group)
    
    def update_calculation_results(self):
        """計算結果ラベルを最新状態に更新"""
        metrics = self._calculate_summary_metrics()
        self.latest_summary_metrics = metrics or {}

        if not metrics:
            self.calculation_label.setText("計算結果: データ未入力")
            return

        def fmt_number(value, digits=0):
            try:
                if value is None:
                    return "0"
                if digits == 0:
                    return f"{float(value):,.0f}"
                return f"{float(value):,.2f}"
            except (ValueError, TypeError):
                return "0"

        text = (
            f"日付: {metrics.get('route_date', '')} / ルート: {metrics.get('route_name', '')}\n"
            f"総仕入点数: {fmt_number(metrics.get('total_item_count'))}点 / "
            f"総仕入額: {fmt_number(metrics.get('total_purchase_amount'))}円 / "
            f"総想定販売額: {fmt_number(metrics.get('total_sales_amount'))}円\n"
            f"総想定粗利: {fmt_number(metrics.get('total_gross_profit'))}円 / "
            f"平均仕入価格: {fmt_number(metrics.get('avg_purchase_price'))}円\n"
            f"総稼働時間: {fmt_number(metrics.get('total_working_hours'), 2)}h / "
            f"想定時給: {fmt_number(metrics.get('estimated_hourly_rate'))}円"
        )
        self.calculation_label.setText(text)
    
    def auto_calculate_store_ratings(self):
        """店舗訪問テーブルの星評価を自動計算"""
        if not hasattr(self, "store_visits_table"):
            return
        row_count = self.store_visits_table.rowCount()
        for row in range(row_count):
            rating = self._calculate_store_rating_for_row(row)
            if rating <= 0:
                rating = 0.0
            else:
                rating = max(0.0, min(5.0, round(rating * 2) / 2))
            star_widget = self.store_visits_table.cellWidget(row, 9)
            if not isinstance(star_widget, StarRatingWidget):
                star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
                star_widget.rating_changed.connect(partial(self.on_star_rating_changed, row))
                self.store_visits_table.setCellWidget(row, 9, star_widget)
            # setRatingでシグナルが発火してUndo履歴が増えないよう一時的にブロック
            block_prev = star_widget.blockSignals(True)
            try:
                star_widget.setRating(rating)
            finally:
                star_widget.blockSignals(block_prev)

    def _calculate_store_rating_for_row(self, row: int) -> float:
        """指定行の星評価を算出"""
        qty = self._parse_float_value(self._get_table_item(row, 8))
        profit = self._parse_float_value(self._get_table_item(row, 7))
        stay = self._parse_float_value(self._get_table_item(row, 5))

        if qty <= 0 or profit <= 0:
            return 0
        stay = max(1.0, stay)

        base_score = self._determine_base_score(int(round(qty)))
        profit_per_minute = profit / stay

        profit_threshold = 190.0
        profit_scale = 30.0
        profit_factor = max(0.0, min(5.0, (profit_per_minute - profit_threshold) / profit_scale))

        base_weight = 0.7
        profit_weight = 0.3

        final_score = (base_weight * base_score) + (profit_weight * profit_factor)
        return max(0.0, min(5.0, final_score))

    def _determine_base_score(self, item_count: int) -> int:
        if item_count >= 10:
            return 5
        if item_count >= 7:
            return 4
        if item_count >= 5:
            return 3
        if item_count >= 3:
            return 2
        return 1

    def _parse_float_value(self, value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return 0.0
        text = text.replace(',', '')
        try:
            return float(text)
        except ValueError:
            return 0.0

    def update_route_codes(self):
        """ルートコード一覧を更新"""
        try:
            # 既存のルート名を取得
            route_names = self.store_db.get_route_names()
            
            # コンボボックスをクリア
            self.route_code_combo.clear()
            
            # ルート名を追加
            for route_name in route_names:
                self.route_code_combo.addItem(route_name)
            
            # 空の選択肢も追加（手動入力用）
            if not self.route_code_combo.findText("") >= 0:
                self.route_code_combo.addItem("")
            
            # デフォルトは空白に設定
            self.route_code_combo.setCurrentText("")
                
        except Exception as e:
            print(f"ルートコード一覧更新エラー: {e}")
    
    def on_route_code_changed(self, route_name: str):
        """ルートコード変更時の処理"""
        try:
            if route_name:
                # ルート名からルートコードを取得
                route_code = self.store_db.get_route_code_by_name(route_name)
                if route_code:
                    # ルートコードを表示用に更新（必要に応じて）
                    pass
        except Exception as e:
            print(f"ルートコード変更処理エラー: {e}")
    
    def get_selected_route_code(self) -> str:
        """選択されたルートコードを取得"""
        route_name = self.route_code_combo.currentText().strip()
        if route_name:
            # ルート名からルートコードを取得
            route_code = self.store_db.get_route_code_by_name(route_name)
            return route_code or route_name  # ルートコードが見つからない場合はルート名を返す
        return ""
    
    def get_stores_from_table(self) -> List[Dict[str, Any]]:
        """テーブルから訪問順序に基づいて店舗一覧を取得"""
        try:
            stores = []
            table = self.store_visits_table
            
            # テーブルの行数
            row_count = table.rowCount()
            if row_count == 0:
                return []
            
            # 各行から訪問順序と店舗コードを取得
            table_data = []
            for row in range(row_count):
                # 訪問順序（0列目）
                visit_order_item = table.item(row, 0)
                visit_order = None
                if visit_order_item:
                    try:
                        visit_order = int(visit_order_item.text())
                    except (ValueError, AttributeError):
                        visit_order = row + 1  # デフォルト値
                else:
                    visit_order = row + 1
                
                # 店舗コード（1列目）
                store_code_item = table.item(row, 1)
                store_code = store_code_item.text().strip() if store_code_item else ''
                
                # 店舗名（2列目）
                store_name_item = table.item(row, 2)
                store_name = store_name_item.text().strip() if store_name_item else ''
                
                if store_code:  # 店舗コードがある行のみ処理
                    table_data.append({
                        'visit_order': visit_order,
                        'store_code': store_code,
                        'store_name': store_name,
                        'row': row
                    })
            
            # 訪問順序でソート
            table_data.sort(key=lambda x: x['visit_order'])
            
            # 店舗マスタから詳細情報を取得
            for data in table_data:
                store_code = data['store_code']
                store_info = self.store_db.get_store_by_code(store_code)
                if store_info:
                    # 店舗マスタの情報をコピー
                    store = dict(store_info)
                    # テーブルの店舗名を優先（ユーザーが変更している可能性があるため）
                    if data['store_name']:
                        store['store_name'] = data['store_name']
                    # 店舗マスタの備考欄を取得
                    custom_fields = store_info.get('custom_fields', {})
                    notes = custom_fields.get('notes', '')
                    store['notes'] = notes
                    stores.append(store)
                else:
                    # 店舗マスタにない場合はテーブルの情報のみで作成
                    store = {
                        'store_code': store_code,
                        'supplier_code': store_code,  # 互換性のため
                        'store_name': data['store_name'],
                        'notes': ''
                    }
                    stores.append(store)
            
            return stores
        except Exception as e:
            print(f"テーブルから店舗一覧取得エラー: {e}")
            import traceback
            print(traceback.format_exc())
            return []
    
    def get_stores_for_route(self, route_name: str) -> List[Dict[str, Any]]:
        """指定されたルートの店舗一覧を取得（表示順序でソート）"""
        try:
            # 表示順序でソートされた店舗一覧を取得
            if hasattr(self.store_db, 'get_stores_for_route_ordered'):
                return self.store_db.get_stores_for_route_ordered(route_name)
            else:
                # 後方互換性のため、表示順序がない場合は従来の方法
                stores = self.store_db.list_stores()
                route_stores = [
                    store for store in stores
                    if store.get('affiliated_route_name') == route_name
                ]
                # display_orderでソート（存在しない場合は0）
                route_stores.sort(key=lambda x: (x.get('display_order', 0), x.get('store_name', '')))
                return route_stores
        except Exception as e:
            print(f"店舗一覧取得エラー: {e}")
            return []
    
    def setup_shortcuts(self):
        """ショートカットキーの設定"""
        # Ctrl+Z: Undo
        undo_shortcut = QShortcut(QKeySequence("Ctrl+Z"), self)
        undo_shortcut.activated.connect(self.undo_action)
        
        # Ctrl+Y: Redo
        redo_shortcut = QShortcut(QKeySequence("Ctrl+Y"), self)
        redo_shortcut.activated.connect(self.redo_action)
    
    def _snapshot_table_state(self):
        """現在のテーブル状態をスナップショットとして返す（星評価含む）"""
        state = []
        for row in range(self.store_visits_table.rowCount()):
            row_data = {}
            for col in range(self.store_visits_table.columnCount()):
                if col == 9:
                    star_widget = self.store_visits_table.cellWidget(row, col)
                    row_data[col] = int(star_widget.rating()) if star_widget else 0
                else:
                    item = self.store_visits_table.item(row, col)
                    row_data[col] = item.text() if item else ''
            state.append(row_data)
        return state

    def save_table_state(self):
        """テーブルの現在の状態を保存（Undo用）"""
        state = self._snapshot_table_state()
        # 直前と同じ状態は保存しない（無駄な履歴を防止）
        if self.undo_stack and self.undo_stack[-1] == state:
            return
        self.undo_stack.append(state)
        if len(self.undo_stack) > self.max_undo_history:
            self.undo_stack.pop(0)
        self.redo_stack.clear()
    
    def restore_table_state(self, state):
        """テーブルの状態を復元"""
        # 変更イベントを一時的に無効化
        self.store_visits_table.blockSignals(True)
        
        try:
            self.store_visits_table.setRowCount(len(state))
            for row_idx, row_data in enumerate(state):
                for col_idx, value in row_data.items():
                    col_idx_int = int(col_idx)
                    # 星評価列の場合はウィジェットを設定
                    if col_idx_int == 9:
                        # self._safe_int が未定義な環境でも安全にパース
                        try:
                            rating = int(float(str(value))) if str(value) not in (None, '', 'None') else 0
                        except Exception:
                            rating = 0
                        star_widget = StarRatingWidget(self.store_visits_table, rating=rating)
                        star_widget.rating_changed.connect(lambda rating, r=row_idx: self.on_star_rating_changed(r, rating))
                        self.store_visits_table.setCellWidget(row_idx, col_idx_int, star_widget)
                    else:
                        item = QTableWidgetItem(str(value))
                        self.store_visits_table.setItem(row_idx, col_idx_int, item)
            
            # 訪問順序を再設定
            self.update_visit_order()
            # 行の高さを内容に合わせて自動調整（折り返しテキスト対応）
            self.store_visits_table.resizeRowsToContents()
            getattr(self, 'update_calculation_results', lambda: None)()
        finally:
            # 変更イベントを再有効化
            self.store_visits_table.blockSignals(False)
    
    def undo_action(self):
        """Undo操作"""
        if not self.undo_stack:
            QMessageBox.information(self, "情報", "元に戻す操作がありません")
            return
        
        # 現在の状態をRedoスタックに保存
        current_state = self._snapshot_table_state()
        self.redo_stack.append(current_state)
        
        # Undoスタックから前の状態を取得
        previous_state = self.undo_stack.pop()
        # 直前に保存された状態が現状態と同一の場合、さらに一つ前を使う
        if previous_state == current_state and self.undo_stack:
            previous_state = self.undo_stack.pop()
        if previous_state != current_state:
            self.restore_table_state(previous_state)
            QMessageBox.information(self, "完了", "操作を元に戻しました")
        else:
            QMessageBox.information(self, "情報", "元に戻す操作がありません")
        # 訪問順序を保存
        self.save_store_order()
    
    def redo_action(self):
        """Redo操作"""
        if not self.redo_stack:
            QMessageBox.information(self, "情報", "やり直す操作がありません")
            return
        
        # 現在の状態をUndoスタックに保存
        current_state = []
        for row in range(self.store_visits_table.rowCount()):
            row_data = {}
            for col in range(self.store_visits_table.columnCount()):
                item = self.store_visits_table.item(row, col)
                row_data[col] = item.text() if item else ''
            current_state.append(row_data)
        self.undo_stack.append(current_state)
        
        # Redoスタックから次の状態を復元
        next_state = self.redo_stack.pop()
        self.restore_table_state(next_state)
        
        # 訪問順序を保存
        self.save_store_order()
        
        QMessageBox.information(self, "完了", "操作をやり直しました")
    
    def on_table_item_changed(self, item):
        """テーブルのアイテムが変更されたときの処理"""
        # 頻繁に呼ばれるので、少し遅延させてから保存（連続変更を1回として扱う）
        if not hasattr(self, '_change_timer'):
            from PySide6.QtCore import QTimer
            self._change_timer = QTimer()
            self._change_timer.setSingleShot(True)
            def _batched_update():
                self.save_table_state()
                self.recalc_travel_times()
            self._change_timer.timeout.connect(_batched_update)
        
        # タイマーをリセット（500ms後に保存）
        self._change_timer.stop()
        self._change_timer.start(500)
    
    def on_rows_moved(self, parent, start, end, destination, row):
        """行移動時の処理（訪問順序を再設定）"""
        # 変更イベントを一時的に無効化
        self.store_visits_table.blockSignals(True)
        
        try:
            # 訪問順序を再設定
            for i in range(self.store_visits_table.rowCount()):
                order_item = self.store_visits_table.item(i, 0)
                if order_item:
                    order_item.setText(str(i + 1))
            
            # 状態を保存
            self.save_table_state()
            
            # 訪問順序を保存
            self.save_store_order()
            getattr(self, 'update_calculation_results', lambda: None)()
        finally:
            # 変更イベントを再有効化
            self.store_visits_table.blockSignals(False)
    
    def save_store_order(self):
        """現在の訪問順序をデータベースに保存（自動保存用）"""
        try:
            route_name = self.route_code_combo.currentText().strip()
            if not route_name:
                return
            
            # 現在の訪問順序を取得
            store_orders = {}
            for row in range(self.store_visits_table.rowCount()):
                code_item = self.store_visits_table.item(row, 1)  # 店舗コード列
                if code_item:
                    supplier_code = code_item.text().strip()
                    if supplier_code:
                        store_orders[supplier_code] = row + 1  # 1始まりの順序
            
            if store_orders:
                # データベースに保存
                if hasattr(self.store_db, 'update_store_display_order'):
                    self.store_db.update_store_display_order(route_name, store_orders)
        except Exception as e:
            print(f"訪問順序保存エラー: {e}")
    
    def save_visit_order_to_db(self):
        """訪問順序をデータベースに保存（ボタンから明示的に呼び出し）"""
        try:
            route_name = self.route_code_combo.currentText().strip()
            if not route_name:
                QMessageBox.warning(self, "警告", "ルートが選択されていません。")
                return
            
            # テーブルにデータがあるか確認
            if self.store_visits_table.rowCount() == 0:
                QMessageBox.warning(self, "警告", "保存する店舗データがありません。")
                return
            
            # 現在の訪問順序を取得
            store_orders = {}
            for row in range(self.store_visits_table.rowCount()):
                code_item = self.store_visits_table.item(row, 1)  # 店舗コード列
                if code_item:
                    store_code = code_item.text().strip()
                    if store_code:
                        store_orders[store_code] = row + 1  # 1始まりの順序
            
            if not store_orders:
                QMessageBox.warning(self, "警告", "保存する店舗コードがありません。")
                return
            
            # データベースに保存
            if hasattr(self.store_db, 'update_store_display_order'):
                success = self.store_db.update_store_display_order(route_name, store_orders)
                if success:
                    QMessageBox.information(
                        self, 
                        "保存完了", 
                        f"訪問順序を保存しました。\n\n"
                        f"ルート: {route_name}\n"
                        f"保存件数: {len(store_orders)}件\n\n"
                        f"次回同じルートを呼び出した時に、保存された順序で表示されます。"
                    )
                else:
                    QMessageBox.warning(self, "エラー", "訪問順序の保存に失敗しました。")
            else:
                QMessageBox.warning(self, "エラー", "データベースに保存機能がありません。")
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"訪問順序保存エラー: {e}")
            print(f"エラー詳細:\n{error_detail}")
            QMessageBox.critical(
                self, 
                "エラー", 
                f"訪問順序の保存中にエラーが発生しました:\n{str(e)}\n\n詳細はコンソールを確認してください。"
            )
    
    def load_template(self) -> Optional[str]:
        """
        テンプレートファイルを読み込む
        
        Returns:
            読み込んだファイルパス（キャンセル時はNone）
        """
        try:
            # 前回選択したフォルダを取得
            last_path = self.settings.value("route_template/last_selected", "")
            default_dir = os.path.dirname(last_path) if last_path and os.path.exists(last_path) else ""
            
            # ファイル選択ダイアログ
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "ルートテンプレートファイルを選択",
                default_dir,
                "Excelファイル (*.xlsx *.xls);;CSVファイル (*.csv);;すべてのファイル (*)"
            )
            
            if not file_path:
                return None
            
            # ファイル拡張子で処理を分岐
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext in ['.xlsx', '.xls']:
                # Excelファイルの読み込み
                self._load_excel_template(file_path)
            elif file_ext == '.csv':
                # CSVファイルの読み込み
                self._load_csv_template(file_path)
            else:
                QMessageBox.warning(self, "警告", "サポートされていないファイル形式です。")
                return None
            
            # 読み込み成功時はパスを保存
            self.last_loaded_template_path = file_path
            self.settings.setValue("route_template/last_selected", file_path)
            
            # 計算結果を更新
            getattr(self, 'update_calculation_results', lambda: None)()
            
            return file_path
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"テンプレートの読み込みに失敗しました:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def _load_excel_template(self, file_path: str):
        """Excelテンプレートファイルを読み込む"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # 店舗訪問詳細シートを探す
            visit_sheet = None
            for sheet_name in wb.sheetnames:
                if '店舗訪問詳細' in sheet_name or '訪問' in sheet_name:
                    visit_sheet = wb[sheet_name]
                    break
            
            if not visit_sheet:
                # デフォルトで最初のシートを使用
                visit_sheet = wb.active
            
            # ルート日付とルートコードを取得（1-2行目）
            route_date = None
            route_code = None
            
            # 1行目: 日付
            date_label = visit_sheet.cell(row=1, column=1).value
            if date_label and ('日付' in str(date_label) or 'date' in str(date_label).lower()):
                date_value = visit_sheet.cell(row=1, column=2).value
                if date_value:
                    if isinstance(date_value, datetime):
                        route_date = date_value.strftime('%Y-%m-%d')
                    else:
                        route_date = str(date_value)
            
            # 2行目: ルート
            route_label = visit_sheet.cell(row=2, column=1).value
            if route_label and ('ルート' in str(route_label) or 'route' in str(route_label).lower()):
                route_value = visit_sheet.cell(row=2, column=2).value
                if route_value:
                    route_code = str(route_value).strip()
            
            # ルート情報を設定
            if route_date:
                try:
                    if isinstance(route_date, str):
                        route_date_obj = datetime.strptime(route_date, '%Y-%m-%d').date()
                    else:
                        route_date_obj = route_date
                    self.route_date_edit.setDate(route_date_obj)
                except Exception:
                    pass
            
            if route_code:
                # ルートコードをコンボボックスに設定
                self.update_route_codes()
                idx = self.route_code_combo.findText(route_code)
                if idx >= 0:
                    self.route_code_combo.setCurrentIndex(idx)
                else:
                    # 見つからない場合は追加
                    self.route_code_combo.addItem(route_code)
                    self.route_code_combo.setCurrentText(route_code)
            
            # 店舗訪問詳細を読み込み（3行目がヘッダー、4行目以降がデータ）
            visits = []
            header_row = 3
            data_start_row = 4
            
            # ヘッダーを取得
            headers = []
            for col in range(1, visit_sheet.max_column + 1):
                header_value = visit_sheet.cell(row=header_row, column=col).value
                if header_value:
                    headers.append(str(header_value).strip())
                else:
                    headers.append('')
            
            # 出発時刻・帰宅時刻・往路高速代・復路高速代を保存する変数
            departure_time = None
            return_time = None
            toll_fee_outbound = 0
            toll_fee_return = 0
            
            # データ行を読み込み
            for row in range(data_start_row, visit_sheet.max_row + 1):
                # 店舗コードが空の行はスキップ
                store_code_cell = visit_sheet.cell(row=row, column=1)
                store_code = store_code_cell.value
                if not store_code or str(store_code).strip() == '':
                    continue
                
                # 除外対象の店舗コードをチェック（値を取得してからスキップ）
                store_code_str = str(store_code).strip()
                if store_code_str == '出発時刻':
                    # 出発時刻を取得（B列）
                    departure_time = self._format_time_value(visit_sheet.cell(row=row, column=2).value)
                    continue
                elif store_code_str == '帰宅時刻':
                    # 帰宅時刻を取得（B列）
                    return_time = self._format_time_value(visit_sheet.cell(row=row, column=2).value)
                    continue
                elif store_code_str == '往路高速代':
                    # 往路高速代を取得（B列）
                    toll_value = visit_sheet.cell(row=row, column=2).value
                    toll_fee_outbound = self._safe_float(str(toll_value or '')) or 0
                    continue
                elif store_code_str == '復路高速代':
                    # 復路高速代を取得（B列）
                    toll_value = visit_sheet.cell(row=row, column=2).value
                    toll_fee_return = self._safe_float(str(toll_value or '')) or 0
                    continue
                
                visit = {
                    'visit_order': len(visits) + 1,
                    'store_code': store_code_str,
                    'store_name': str(visit_sheet.cell(row=row, column=2).value or '').strip(),
                    'store_in_time': self._format_time_value(visit_sheet.cell(row=row, column=3).value),
                    'store_out_time': self._format_time_value(visit_sheet.cell(row=row, column=4).value),
                    'stay_duration': self._safe_float(str(visit_sheet.cell(row=row, column=5).value or '')) or 0.0,
                    'store_notes': str(visit_sheet.cell(row=row, column=6).value or '').strip(),
                }
                visits.append(visit)
            
            # 読み込んだ値をroute_dataに保存
            if departure_time or return_time or toll_fee_outbound or toll_fee_return:
                if not hasattr(self, 'route_data') or not self.route_data:
                    self.route_data = {}
                
                if departure_time:
                    # ルート日付と結合してdatetime形式に
                    if route_date:
                        self.route_data['departure_time'] = f"{route_date} {departure_time}:00"
                    else:
                        self.route_data['departure_time'] = departure_time
                
                if return_time:
                    # ルート日付と結合してdatetime形式に
                    if route_date:
                        self.route_data['return_time'] = f"{route_date} {return_time}:00"
                    else:
                        self.route_data['return_time'] = return_time
                
                if toll_fee_outbound:
                    self.route_data['toll_fee_outbound'] = toll_fee_outbound
                
                if toll_fee_return:
                    self.route_data['toll_fee_return'] = toll_fee_return
            
            # テーブルに反映
            self.store_visits_table.blockSignals(True)
            try:
                self.store_visits_table.setRowCount(len(visits))
                for i, visit in enumerate(visits):
                    # 訪問順序
                    order_item = QTableWidgetItem(str(i + 1))
                    self.store_visits_table.setItem(i, 0, order_item)
                    
                    # 店舗コード
                    self.store_visits_table.setItem(i, 1, QTableWidgetItem(visit.get('store_code', '')))
                    
                    # 店舗名
                    self.store_visits_table.setItem(i, 2, QTableWidgetItem(visit.get('store_name', '')))
                    
                    # IN時間
                    in_time = visit.get('store_in_time', '')
                    self.store_visits_table.setItem(i, 3, QTableWidgetItem(in_time))
                    
                    # OUT時間
                    out_time = visit.get('store_out_time', '')
                    self.store_visits_table.setItem(i, 4, QTableWidgetItem(out_time))
                    
                    # 滞在時間（自動計算されるので空欄）
                    self.store_visits_table.setItem(i, 5, QTableWidgetItem(''))
                    
                    # 移動時間（自動計算されるので空欄）
                    self.store_visits_table.setItem(i, 6, QTableWidgetItem(''))
                    
                    # 想定粗利
                    self.store_visits_table.setItem(i, 7, QTableWidgetItem('0'))
                    
                    # 仕入れ点数
                    self.store_visits_table.setItem(i, 8, QTableWidgetItem('0'))
                    
                    # 評価（星評価ウィジェット）
                    star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
                    star_widget.rating_changed.connect(lambda rating, r=i: self.on_star_rating_changed(r, rating))
                    self.store_visits_table.setCellWidget(i, 9, star_widget)
                    
                    # メモ
                    self.store_visits_table.setItem(i, 10, QTableWidgetItem(visit.get('store_notes', '')))
            finally:
                self.store_visits_table.blockSignals(False)
            
            # 滞在時間と移動時間を自動計算
            self.recalc_travel_times()
            
            # 訪問順序を更新
            self.update_visit_order()
            
            QMessageBox.information(self, "完了", f"テンプレートを読み込みました。\n店舗数: {len(visits)}件")
            
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みエラー: {str(e)}")
    
    def _load_csv_template(self, file_path: str):
        """CSVテンプレートファイルを読み込む"""
        try:
            # CSVファイルの読み込み（UTF-8 BOM対応）
            df = pd.read_csv(file_path, encoding='utf-8-sig', dtype=str, keep_default_na=False)
            
            # ルート情報セクションと店舗訪問詳細セクションを分離
            route_data = {}
            visits = []
            
            in_route_section = False
            in_visit_section = False
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()
            
            for line in lines:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if 'ルート情報' in line:
                    in_route_section = True
                    in_visit_section = False
                    continue
                elif '店舗訪問詳細' in line:
                    in_route_section = False
                    in_visit_section = True
                    continue
                
                if in_route_section:
                    # ルート情報の処理（簡易実装）
                    parts = line.split(',')
                    if len(parts) >= 2:
                        key = parts[0].strip()
                        value = parts[1].strip() if len(parts) > 1 else ''
                        if key == 'ルート日付':
                            route_data['route_date'] = value
                        elif key == 'ルートコード':
                            route_data['route_code'] = value
                        elif key == '出発時間':
                            route_data['departure_time'] = value
                        elif key == '帰宅時間':
                            route_data['return_time'] = value
                        elif key == '往路高速代':
                            try:
                                route_data['toll_fee_outbound'] = float(value) if value else 0
                            except (ValueError, TypeError):
                                route_data['toll_fee_outbound'] = 0
                        elif key == '復路高速代':
                            try:
                                route_data['toll_fee_return'] = float(value) if value else 0
                            except (ValueError, TypeError):
                                route_data['toll_fee_return'] = 0
                elif in_visit_section:
                    # 店舗訪問詳細の処理
                    parts = line.split(',')
                    if len(parts) >= 2 and parts[0].strip() != '訪問順序':
                        visit = {
                            'visit_order': len(visits) + 1,
                            'store_code': parts[1].strip() if len(parts) > 1 else '',
                            'store_name': parts[2].strip() if len(parts) > 2 else '',
                            'store_in_time': parts[3].strip() if len(parts) > 3 else '',
                            'store_out_time': parts[4].strip() if len(parts) > 4 else '',
                            'store_notes': parts[11].strip() if len(parts) > 11 else '',
                        }
                        if visit['store_code']:
                            visits.append(visit)
            
            # ルート情報を設定
            route_date_str = None
            if 'route_date' in route_data:
                route_date_str = route_data['route_date']
                try:
                    route_date_obj = datetime.strptime(route_date_str, '%Y-%m-%d').date()
                    self.route_date_edit.setDate(route_date_obj)
                except Exception:
                    pass
            
            if 'route_code' in route_data:
                self.update_route_codes()
                route_code = route_data['route_code']
                idx = self.route_code_combo.findText(route_code)
                if idx >= 0:
                    self.route_code_combo.setCurrentIndex(idx)
                else:
                    self.route_code_combo.addItem(route_code)
                    self.route_code_combo.setCurrentText(route_code)
            
            # 出発時刻・帰宅時刻・往路高速代・復路高速代をroute_dataに保存
            if not hasattr(self, 'route_data') or not self.route_data:
                self.route_data = {}
            
            if 'departure_time' in route_data:
                dep_time = route_data['departure_time']
                if route_date_str:
                    self.route_data['departure_time'] = f"{route_date_str} {dep_time}:00" if ' ' not in str(dep_time) else dep_time
                else:
                    self.route_data['departure_time'] = dep_time
            
            if 'return_time' in route_data:
                ret_time = route_data['return_time']
                if route_date_str:
                    self.route_data['return_time'] = f"{route_date_str} {ret_time}:00" if ' ' not in str(ret_time) else ret_time
                else:
                    self.route_data['return_time'] = ret_time
            
            if 'toll_fee_outbound' in route_data:
                self.route_data['toll_fee_outbound'] = route_data['toll_fee_outbound']
            
            if 'toll_fee_return' in route_data:
                self.route_data['toll_fee_return'] = route_data['toll_fee_return']
            
            # テーブルに反映（Excelと同じ処理）
            self.store_visits_table.blockSignals(True)
            try:
                self.store_visits_table.setRowCount(len(visits))
                for i, visit in enumerate(visits):
                    order_item = QTableWidgetItem(str(i + 1))
                    self.store_visits_table.setItem(i, 0, order_item)
                    self.store_visits_table.setItem(i, 1, QTableWidgetItem(visit.get('store_code', '')))
                    self.store_visits_table.setItem(i, 2, QTableWidgetItem(visit.get('store_name', '')))
                    self.store_visits_table.setItem(i, 3, QTableWidgetItem(visit.get('store_in_time', '')))
                    self.store_visits_table.setItem(i, 4, QTableWidgetItem(visit.get('store_out_time', '')))
                    self.store_visits_table.setItem(i, 5, QTableWidgetItem(''))
                    self.store_visits_table.setItem(i, 6, QTableWidgetItem(''))
                    self.store_visits_table.setItem(i, 7, QTableWidgetItem('0'))
                    self.store_visits_table.setItem(i, 8, QTableWidgetItem('0'))
                    star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
                    star_widget.rating_changed.connect(lambda rating, r=i: self.on_star_rating_changed(r, rating))
                    self.store_visits_table.setCellWidget(i, 9, star_widget)
                    self.store_visits_table.setItem(i, 10, QTableWidgetItem(visit.get('store_notes', '')))
            finally:
                self.store_visits_table.blockSignals(False)
            
            self.recalc_travel_times()
            self.update_visit_order()
            
            QMessageBox.information(self, "完了", f"テンプレートを読み込みました。\n店舗数: {len(visits)}件")
            
        except Exception as e:
            raise Exception(f"CSVファイルの読み込みエラー: {str(e)}")
    
    def _format_time_value(self, value) -> str:
        """時刻値を文字列に変換"""
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%H:%M')
        if isinstance(value, dt_time):
            return value.strftime('%H:%M')
        value_str = str(value).strip()
        if ':' in value_str:
            # HH:MM形式の場合はそのまま返す
            parts = value_str.split(':')
            if len(parts) >= 2:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
        return value_str
    
    def generate_template(self):
        """テンプレート生成"""
        try:
            route_name = self.route_code_combo.currentText().strip()
            route_code = self.get_selected_route_code()
            
            # デフォルトフォルダを取得
            default_dir = self.template_save_default_dir if self.template_save_default_dir and os.path.isdir(self.template_save_default_dir) else ""
            default_filename = f"route_template_{route_name or 'new'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            default_path = os.path.join(default_dir, default_filename) if default_dir else default_filename
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "テンプレートファイルを保存",
                default_path,
                "Excelファイル (*.xlsx);;CSVファイル (*.csv)"
            )
            
            # 保存先フォルダを設定に保存
            if file_path:
                save_dir = os.path.dirname(file_path)
                if save_dir and os.path.isdir(save_dir):
                    self.update_template_save_default_dir(save_dir)
            
            if not file_path:
                return None
            
            # 選択されたルートの店舗一覧を取得
            stores = []
            store_codes = []
            if route_name:
                # テーブルから訪問順序を取得（テーブルにデータがある場合）
                table_stores = self.get_stores_from_table()
                if table_stores:
                    # テーブルの訪問順序を使用
                    stores = table_stores
                    store_codes = [
                        (store.get('store_code') or store.get('supplier_code'))
                        for store in stores
                        if store.get('store_code') or store.get('supplier_code')
                    ]
                else:
                    # テーブルにデータがない場合はデータベースから取得
                    stores = self.get_stores_for_route(route_name)
                    # 店舗マスタの備考欄を取得してstoresに追加
                    for store in stores:
                        # 店舗コード(store_code)を優先し、互換性のため仕入れ先コードも許容
                        any_code = store.get('store_code') or store.get('supplier_code')
                        if any_code:
                            store_info = self.store_db.get_store_by_code(any_code)
                            if store_info:
                                custom_fields = store_info.get('custom_fields', {})
                                notes = custom_fields.get('notes', '')
                                # storesに備考を追加（テンプレート生成時に使用）
                                store['notes'] = notes
                    store_codes = [
                        (store.get('store_code') or store.get('supplier_code'))
                        for store in stores
                        if store.get('store_code') or store.get('supplier_code')
                    ]
            
            if not TemplateGenerator:
                QMessageBox.warning(self, "エラー", "テンプレート生成機能が利用できません")
                return
            
            # ルート日付を取得（QDateをdatetime.dateに変換）
            try:
                qdate = self.route_date_edit.dateTime().date()
                # QDateをdatetime.dateに変換
                route_date = datetime(qdate.year(), qdate.month(), qdate.day()).date()
                print(f"ルート日付取得: {route_date} (QDate: {qdate.year()}-{qdate.month()}-{qdate.day()})")
            except Exception as e:
                print(f"ルート日付取得エラー: {e}")
                import traceback
                print(traceback.format_exc())
                route_date = None
            
            if file_path.endswith('.xlsx'):
                # ルート名（日本語名）とルート日付を渡してテンプレートに表示
                print(f"Excelテンプレート生成開始: ファイル={file_path}, ルート名={route_name}, 店舗数={len(stores) if stores else len(store_codes) if store_codes else 0}, ルート日付={route_date}")
                success = TemplateGenerator.generate_excel_template(file_path, route_name, store_codes, stores, route_date)
            else:
                # CSVテンプレートにもルート日付を渡す
                print(f"CSVテンプレート生成開始: ファイル={file_path}, ルートコード={route_code}, 店舗数={len(store_codes) if store_codes else 0}, ルート日付={route_date}")
                success = TemplateGenerator.generate_csv_template(file_path, route_code, store_codes, route_date)
            
            if success:
                QMessageBox.information(self, "成功", f"テンプレートを生成しました:\n{file_path}")
            else:
                QMessageBox.warning(self, "エラー", "テンプレートの生成に失敗しました。\n詳細はコンソールを確認してください。")
                
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"テンプレート生成エラー詳細:\n{error_detail}")
            QMessageBox.critical(self, "エラー", f"テンプレート生成中にエラーが発生しました:\n{str(e)}\n\n詳細はコンソールを確認してください。")
    
    def browse_template_save_dir(self):
        """テンプレート保存フォルダの選択"""
        current_dir = self.template_save_default_dir if self.template_save_default_dir and os.path.isdir(self.template_save_default_dir) else str(Path.home())
        selected_dir = QFileDialog.getExistingDirectory(
            self,
            "テンプレート保存フォルダを選択",
            current_dir
        )
        if selected_dir:
            self.update_template_save_default_dir(selected_dir)
    
    def on_template_save_dir_edit_finished(self):
        """手入力でテンプレート保存フォルダを更新"""
        text = self.template_save_dir_edit.text().strip()
        if not text:
            self.update_template_save_default_dir("")
            return
        expanded = os.path.expanduser(text)
        if not os.path.isdir(expanded):
            QMessageBox.warning(self, "エラー", "指定したフォルダが存在しません。")
            # 元の値へ戻す
            self.template_save_dir_edit.blockSignals(True)
            self.template_save_dir_edit.setText(self.template_save_default_dir)
            self.template_save_dir_edit.blockSignals(False)
            return
        self.update_template_save_default_dir(os.path.abspath(expanded))
    
    def update_template_save_default_dir(self, directory: str):
        """テンプレート保存フォルダの設定を保存"""
        normalized = directory if directory else ""
        if normalized and not os.path.isdir(normalized):
            return
        self.template_save_default_dir = normalized
        if hasattr(self, "template_save_dir_edit"):
            self.template_save_dir_edit.blockSignals(True)
            self.template_save_dir_edit.setText(self.template_save_default_dir)
            self.template_save_dir_edit.blockSignals(False)
        if self.settings is not None:
            self.settings.setValue("route_template/default_save_dir", self.template_save_default_dir)
    
    def run_matching(self):
        """照合処理実行（改良版：仕入管理データを参照）"""
        try:
            # 現在のルートIDが必要。未保存なら保存を促し、自動保存を試みる
            if not self.current_route_id:
                reply = QMessageBox.question(
                    self,
                    "未保存のルート",
                    "照合処理を実行するにはルートの保存が必要です。今すぐ保存しますか？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                if reply == QMessageBox.Yes:
                    self.save_data()
                if not self.current_route_id:
                    QMessageBox.warning(self, "警告", "ルートIDが未確定のため処理を中止しました")
                    return
            
            # 仕入管理データの確認
            if not self.inventory_widget:
                QMessageBox.warning(self, "警告", "仕入管理ウィジェットへの参照がありません")
                return
            
            inventory_data = self.inventory_widget.inventory_data
            if inventory_data is None or len(inventory_data) == 0:
                # データがない場合、CSVファイル選択にフォールバック
                reply = QMessageBox.question(
                    self,
                    "データなし",
                    "仕入管理にデータがありません。\nCSVファイルを選択して処理しますか？",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    # 既存のCSVファイル選択処理を呼び出し
                    self.execute_matching_from_csv()
                return
            
            # 時間許容誤差の設定
            tolerance, ok = QInputDialog.getInt(
                self,
                "時間許容誤差",
                "時間許容誤差（分）:",
                30, 0, 120, 1
            )
            if not ok:
                return
            
            # データをJSON形式に変換（NaN値を事前に処理）
            # NaN値を空文字列に置換してからJSON化
            clean_data = inventory_data.fillna('')
            purchase_data = clean_data.to_dict(orient="records")
            
            # API呼び出し
            QMessageBox.information(self, "処理中", "照合処理を実行しています...")
            result = self.api_client.inventory_match_stores_from_data(
                purchase_data=purchase_data,
                route_summary_id=self.current_route_id,
                time_tolerance_minutes=tolerance
            )
            
            # 結果を仕入管理ウィジェットに反映
            if result.get('status') == 'success':
                # 照合後のデータで仕入管理データを更新
                result_data = result.get('data', [])
                if result_data:
                    import pandas as pd
                    updated_df = pd.DataFrame(result_data)
                    
                    # 仕入管理ウィジェットのデータを更新
                    self.inventory_widget.inventory_data = updated_df
                    self.inventory_widget.filtered_data = updated_df.copy()
                    self.inventory_widget.update_table()
                    self.inventory_widget.update_data_count()
                
                # 店舗コード別の粗利を集計してルートサマリーを更新
                self._update_route_gross_profit_from_inventory(result_data)
                
                # 結果表示
                stats = result.get('stats', {})
                matched_rows = stats.get('matched_rows', 0)
                total_rows = stats.get('total_rows', 0)
                
                msg = f"照合処理完了\n\n総行数: {total_rows}\nマッチした行数: {matched_rows}"
                msg += "\n\n仕入管理タブのデータが更新され、\nルートサマリーの想定粗利も自動計算されました。"
                QMessageBox.information(self, "照合処理完了", msg)
                self.update_calculation_results()
            else:
                QMessageBox.warning(self, "エラー", "照合処理に失敗しました")
                
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"照合処理中にエラーが発生しました:\n{str(e)}")
    
    def execute_matching_from_csv(self):
        """照合処理実行（CSVファイル版）"""
        try:
            # 現在のルートIDが必要。未保存なら保存を促し、自動保存を試みる
            if not self.current_route_id:
                reply = QMessageBox.question(
                    self,
                    "未保存のルート",
                    "照合処理を実行するにはルートの保存が必要です。今すぐ保存しますか？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                if reply == QMessageBox.Yes:
                    self.save_data()
                if not self.current_route_id:
                    QMessageBox.warning(self, "警告", "ルートIDが未確定のため処理を中止しました")
                    return
            
            # APIクライアント確認
            if not self.api_client:
                from api.client import APIClient
                self.api_client = APIClient()
            
            # CSVファイル選択
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "仕入CSVファイルを選択",
                "",
                "CSVファイル (*.csv);;すべてのファイル (*)"
            )
            if not file_path:
                return
            
            # 時間許容誤差の設定（デフォルト30分）
            tolerance, ok = QInputDialog.getInt(
                self,
                "時間許容誤差",
                "時間許容誤差（分）:",
                30, 0, 120, 1
            )
            if not ok:
                return
            
            # API呼び出し
            QMessageBox.information(self, "処理中", "照合処理を実行しています...")
            result = self.api_client.inventory_match_stores(
                file_path=file_path,
                route_summary_id=self.current_route_id,
                time_tolerance_minutes=tolerance
            )
            
            # 結果表示
            stats = result.get('stats', {})
            matched_rows = stats.get('matched_rows', 0)
            total_rows = stats.get('total_rows', 0)
            
            msg = f"照合処理完了\n\n総行数: {total_rows}\nマッチした行数: {matched_rows}"
            if matched_rows > 0:
                msg += f"\n\nプレビュー（先頭10件）には店舗コードが自動付与されています。"
            
            QMessageBox.information(self, "照合処理完了", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"照合処理中にエラーが発生しました:\n{str(e)}")
    
    def load_saved_data(self, route_id: int):
        """指定IDの保存データを読み込む"""
        try:
            row = self.route_db.get_route_summary(route_id)
            if not row:
                QMessageBox.information(self, "情報", "保存済みデータが見つかりませんでした")
                return
            self.current_route_id = route_id
            # 上部フィールド
            self.route_date_edit.setDateTime(QDateTime.fromString(row.get('route_date',''), 'yyyy-MM-dd'))
            
            # ルートコードを読み込んだルートに更新（日本語名で表示）
            route_code = row.get('route_code', '')
            
            if route_code:
                # シグナルを一時的にブロック（currentTextChangedが発火しないように）
                self.route_code_combo.blockSignals(True)
                try:
                    # コンボボックスのアイテムを先に更新（最新のルート一覧を反映）
                    self.update_route_codes()
                    
                    # ルートコードから日本語名を取得
                    route_name = self.store_db.get_route_name_by_code(route_code)
                    print(f"\n=== 保存データ読み込みデバッグ ===")
                    print(f"DBから取得したroute_code: '{route_code}'")
                    print(f"get_route_name_by_code('{route_code}') の結果: '{route_name}'")
                    
                    # もし取得できなかった場合、ルートコードが既に日本語名の可能性がある
                    if not route_name:
                        # ルートコードがそのまま日本語名の場合（後方互換性のため）
                        # コンボボックスにその値があるかチェック
                        if self.route_code_combo.findText(route_code) >= 0:
                            route_name = route_code
                            print(f"route_codeが直接日本語名として存在: '{route_code}'")
                    
                    display_value = route_name if route_name else route_code
                    print(f"表示値（display_value）: '{display_value}'")
                    
                    # コンボボックスに該当するアイテムがあるかチェック
                    idx = self.route_code_combo.findText(display_value)
                    print(f"findText('{display_value}') の結果: idx={idx}")
                    
                    if idx >= 0:
                        # アイテムが見つかった場合は選択
                        self.route_code_combo.setCurrentIndex(idx)
                        print(f"setCurrentIndex({idx}) を実行")
                    else:
                        # アイテムが見つからない場合は追加してから選択
                        print(f"アイテムが見つからないため追加します")
                        self.route_code_combo.addItem(display_value)
                        idx = self.route_code_combo.findText(display_value)
                        if idx >= 0:
                            self.route_code_combo.setCurrentIndex(idx)
                            print(f"追加後に setCurrentIndex({idx}) を実行")
                        else:
                            # 最後の手段：直接テキストを設定（編集可能なので）
                            self.route_code_combo.setCurrentText(display_value)
                            print(f"直接 setCurrentText('{display_value}') を実行")
                    
                    # 最終的な表示値を確認
                    final_display = self.route_code_combo.currentText()
                    print(f"最終的な表示値: '{final_display}'")
                    print(f"==============================\n")
                finally:
                    # シグナルのブロックを解除
                    self.route_code_combo.blockSignals(False)
            else:
                # ルートコードが空の場合は、コンボボックスを更新してクリア
                self.update_route_codes()
                self.route_code_combo.setCurrentText('')
            # 出発時間・帰宅時間・経費・備考は削除されたため、読み込み処理をスキップ
            
            # 店舗訪問詳細
            visits = self.route_db.get_store_visits_by_route(route_id)
            # ルート名からコード→店名の簡易マップを作成（補完用）
            code_to_name = {}
            try:
                route_name = self.route_code_combo.currentText().strip()
                if route_name:
                    for s in self.get_stores_for_route(route_name):
                        c = s.get('store_code') or s.get('supplier_code')
                        n = s.get('store_name')
                        if c and n:
                            code_to_name[c] = n
            except Exception:
                pass
            self.store_visits_table.blockSignals(True)
            try:
                self.store_visits_table.setRowCount(len(visits))
                for i, v in enumerate(visits):
                    order_item = QTableWidgetItem(str(i + 1))
                    order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                    self.store_visits_table.setItem(i, 0, order_item)
                    self.store_visits_table.setItem(i, 1, QTableWidgetItem(v.get('store_code','')))
                    # store_name はDBのstore_visit_detailsに無いのでマスタから補完
                    code = v.get('store_code','')
                    store_name = code_to_name.get(code, '')
                    if not store_name and code:
                        try:
                            # 店舗コード(store_code)を優先し、互換性のため仕入れ先コードも許容
                            s = self.store_db.get_store_by_code(code)
                            store_name = (s or {}).get('store_name','')
                        except Exception:
                            store_name = ''
                    self.store_visits_table.setItem(i, 2, QTableWidgetItem(store_name))
                    # 店舗IN/OUT時間は "2025-10-26 17:22:00" 形式から HH:MM を抽出
                    store_in_time = v.get('store_in_time') or ''
                    store_out_time = v.get('store_out_time') or ''
                    try:
                        store_in_display = store_in_time.split(' ')[1][:5] if ' ' in store_in_time else store_in_time[:5]
                    except Exception:
                        store_in_display = ''
                    try:
                        store_out_display = store_out_time.split(' ')[1][:5] if ' ' in store_out_time else store_out_time[:5]
                    except Exception:
                        store_out_display = ''
                    self.store_visits_table.setItem(i, 3, QTableWidgetItem(store_in_display))
                    self.store_visits_table.setItem(i, 4, QTableWidgetItem(store_out_display))
                    self.store_visits_table.setItem(i, 5, QTableWidgetItem(str(v.get('stay_duration') or '')))
                    self.store_visits_table.setItem(i, 6, QTableWidgetItem(str(v.get('travel_time_from_prev') or '')))
                    # 想定粗利は整数で表示（None/空の場合は0）
                    gross_profit = v.get('store_gross_profit')
                    if gross_profit is None or gross_profit == '':
                        self.store_visits_table.setItem(i, 7, QTableWidgetItem('0'))
                    else:
                        self.store_visits_table.setItem(i, 7, QTableWidgetItem(str(int(float(gross_profit)))))
                    # 仕入れ点数は整数で表示（None/空の場合は0）
                    item_count = v.get('store_item_count')
                    if item_count is None or item_count == '':
                        self.store_visits_table.setItem(i, 8, QTableWidgetItem('0'))
                    else:
                        self.store_visits_table.setItem(i, 8, QTableWidgetItem(str(int(item_count))))
                    existing_rating = v.get('store_rating')
                    try:
                        rating_value = float(existing_rating) if existing_rating not in (None, '') else 0.0
                    except (TypeError, ValueError):
                        rating_value = 0.0
                    star_widget = StarRatingWidget(self.store_visits_table, rating=rating_value, star_size=14)
                    star_widget.rating_changed.connect(lambda rating, r=i: self.on_star_rating_changed(r, rating))
                    self.store_visits_table.setCellWidget(i, 9, star_widget)
                    self.store_visits_table.setItem(i, 10, QTableWidgetItem(v.get('store_notes','') or ''))
            finally:
                self.store_visits_table.blockSignals(False)
            
            self.auto_calculate_store_ratings()
            self.update_visit_order()
            getattr(self, 'update_calculation_results', lambda: None)()
            # Undo基点
            self.undo_stack.clear(); self.redo_stack.clear(); self.save_table_state()
            QMessageBox.information(self, "完了", "保存済みデータを読み込みました")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"保存済みデータの読み込みに失敗しました:\n{str(e)}")

    def delete_saved_data(self, route_id: int):
        """指定IDの保存済みデータを削除"""
        try:
            target_id = route_id
            reply = QMessageBox.question(
                self,
                "削除確認",
                "この保存データを削除しますか？（店舗訪問詳細も削除されます）",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

            ok = self.route_db.delete_route_summary(int(target_id))
            if ok:
                # 画面側はクリア
                if self.current_route_id == target_id:
                    self.current_route_id = None
                self.store_visits_table.setRowCount(0)
                self.undo_stack.clear(); self.redo_stack.clear(); self.save_table_state()
                QMessageBox.information(self, "完了", "保存データを削除しました")
            else:
                QMessageBox.warning(self, "警告", "削除できませんでした")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"保存データの削除に失敗しました:\n{str(e)}")

    def open_saved_history(self):
        """保存履歴ダイアログを開き、読み込み/削除を実行"""
        dlg = SavedRoutesDialog(self.route_db, self.store_db, self)
        res = dlg.exec()
        if res == QDialog.Accepted:
            action, route_id = dlg.get_result()
            if action == 'load' and route_id:
                self.load_saved_data(route_id)
            elif action == 'delete' and route_id:
                self.delete_saved_data(route_id)
    def auto_add_stores(self):
        """選択されたルートの店舗を自動追加（重複チェック付き）"""
        # 変更前の状態を保存
        self.save_table_state()
        
        # 店舗マスタから「店舗コード」として使う値を取得するヘルパー
        # 優先: 店舗コード（store_code） → フォールバック: 仕入れ先コード
        def _get_visit_store_code(store: dict) -> str:
            # 1. 店舗コードを優先
            code = (store.get('store_code') or '').strip()
            if code:
                return code
            # 2. 店舗コードが空の場合だけ仕入れ先コードを使用
            return (store.get('supplier_code') or '').strip()
        
        try:
            route_name = self.route_code_combo.currentText().strip()
            if not route_name:
                QMessageBox.warning(self, "警告", "ルートを選択してください")
                return
            
            # 選択されたルートの店舗一覧を取得
            stores = self.get_stores_for_route(route_name)
            if not stores:
                QMessageBox.information(self, "情報", f"ルート「{route_name}」に登録されている店舗がありません")
                return
            
            # 既存の店舗コード一覧を取得（重複チェック用）
            existing_codes = set()
            for row in range(self.store_visits_table.rowCount()):
                code_item = self.store_visits_table.item(row, 1)  # 店舗コード列
                if code_item:
                    code = code_item.text().strip()
                    if code:
                        existing_codes.add(code)
            
            # 追加する店舗をフィルタリング（重複を除外）
            stores_to_add = [
                store for store in stores
                if _get_visit_store_code(store) and _get_visit_store_code(store) not in existing_codes
            ]
            
            if not stores_to_add:
                QMessageBox.information(self, "情報", "追加可能な店舗がありません（すべて既に追加済みです）")
                return
            
            # 既存の行数を取得
            current_rows = self.store_visits_table.rowCount()
            
            # 店舗をテーブルに追加
            for i, store in enumerate(stores_to_add):
                row = current_rows + i
                self.store_visits_table.insertRow(row)
                
                # 訪問順序（編集可能）
                order_item = QTableWidgetItem(str(row + 1))
                self.store_visits_table.setItem(row, 0, order_item)
                
                # 店舗コード（店舗マスタの「仕入れ先コード」を優先し、なければ店舗コードを使用）
                code_item = QTableWidgetItem(_get_visit_store_code(store))
                self.store_visits_table.setItem(row, 1, code_item)
                
                # 店舗名
                name_item = QTableWidgetItem(store.get('store_name', ''))
                self.store_visits_table.setItem(row, 2, name_item)
                
                # 星評価ウィジェットをセルに配置
                star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
                star_widget.rating_changed.connect(lambda rating, r=row: self.on_star_rating_changed(r, rating))
                self.store_visits_table.setCellWidget(row, 9, star_widget)
                
                # 備考（店舗マスタの備考欄から取得）
                notes = store.get('notes', '') or ''
                notes_item = QTableWidgetItem(notes)
                notes_item.setToolTip(notes)  # ツールチップで全文表示
                self.store_visits_table.setItem(row, 10, notes_item)
            
            # 訪問順序を再設定
            self.update_visit_order()
            
            # 訪問順序を保存
            self.save_store_order()
            
            # 行の高さを内容に合わせて自動調整（折り返しテキスト対応）
            self.store_visits_table.resizeRowsToContents()
            
            # 変更後の状態を保存
            self.save_table_state()
            
            QMessageBox.information(self, "完了", f"{len(stores_to_add)}件の店舗を追加しました")
            getattr(self, 'update_calculation_results', lambda: None)()
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"店舗自動追加中にエラーが発生しました:\n{str(e)}")
    
    def update_visit_order(self):
        """訪問順序を再設定"""
        for i in range(self.store_visits_table.rowCount()):
            order_item = self.store_visits_table.item(i, 0)
            if order_item:
                order_item.setText(str(i + 1))
        # 順序が変わったら移動時間も更新
        self.recalc_travel_times()

    def recalc_travel_times(self):
        """滞在時間と移動時間（分）を自動計算して5/6列目に反映"""
        try:
            row_count = self.store_visits_table.rowCount()
            if row_count == 0:
                return

            def parse_hhmm(text: str) -> Optional[QTime]:
                text = (text or '').strip()
                if not text:
                    return None
                # HH:mm または H:mm を許容
                parts = text.split(':')
                if len(parts) != 2:
                    return None
                try:
                    h = int(parts[0])
                    m = int(parts[1])
                    if 0 <= h < 24 and 0 <= m < 60:
                        return QTime(h, m)
                except ValueError:
                    return None
                return None

            def get_text(row: int, col: int) -> str:
                item = self.store_visits_table.item(row, col)
                return item.text() if item else ''

            # 各行の滞在時間（店舗OUT - 店舗IN）を算出
            for r in range(row_count):
                in_t = parse_hhmm(get_text(r, 3))
                out_t = parse_hhmm(get_text(r, 4))
                if in_t and out_t:
                    mins = in_t.secsTo(out_t) // 60
                    mins = max(0, int(mins))
                    self.store_visits_table.setItem(r, 5, QTableWidgetItem(str(mins)))
                else:
                    self.store_visits_table.setItem(r, 5, QTableWidgetItem(''))

            # 1店舗目の移動時間は空（出発時間が削除されたため）
            if row_count > 0:
                self.store_visits_table.setItem(0, 6, QTableWidgetItem(''))

            # 2店舗目以降: 前店舗OUT → 現在IN
            for r in range(1, row_count):
                prev_out = parse_hhmm(get_text(r - 1, 4))
                cur_in = parse_hhmm(get_text(r, 3))
                if prev_out and cur_in:
                    mins = prev_out.secsTo(cur_in) // 60
                    mins = max(0, int(mins))
                    self.store_visits_table.setItem(r, 6, QTableWidgetItem(str(mins)))
                else:
                    self.store_visits_table.setItem(r, 6, QTableWidgetItem(''))
        except Exception as e:
            print(f"移動時間計算エラー: {e}")

    def reorder_by_visit_order(self):
        """訪問順序列に入力された数字に基づいて行を並び替える"""
        try:
            row_count = self.store_visits_table.rowCount()
            if row_count == 0:
                QMessageBox.information(self, "情報", "並び替える行がありません")
                return
            
            # 変更前の状態を保存（Undo用）
            self.save_table_state()
            
            # 各行のデータと訪問順序を取得
            rows_data = []
            for row in range(row_count):
                # 訪問順序を取得
                order_item = self.store_visits_table.item(row, 0)
                order_text = order_item.text().strip() if order_item else ""
                
                # 訪問順序を数値に変換（数値でない場合は元の行番号+1000で最後に配置）
                try:
                    order_num = int(order_text) if order_text else row + 1000
                except ValueError:
                    order_num = row + 1000  # 数値変換できない場合は最後尾へ
                
                # 行データを収集
                row_data = {
                    'order': order_num,
                    'original_row': row,
                    'cells': []
                }
                
                # 全列のデータを保存
                for col in range(self.store_visits_table.columnCount()):
                    item = self.store_visits_table.item(row, col)
                    cell_text = item.text() if item else ""
                    row_data['cells'].append(cell_text)
                
                # 星評価ウィジェットの値も保存
                star_widget = self.store_visits_table.cellWidget(row, 9)
                if star_widget and hasattr(star_widget, 'rating'):
                    row_data['star_rating'] = star_widget.rating()  # メソッド呼び出し
                else:
                    row_data['star_rating'] = 0
                
                rows_data.append(row_data)
            
            # 訪問順序でソート
            rows_data.sort(key=lambda x: x['order'])
            
            # テーブルを一旦ブロック
            self.store_visits_table.blockSignals(True)
            
            try:
                # テーブルをクリアして再構築
                self.store_visits_table.setRowCount(0)
                self.store_visits_table.setRowCount(len(rows_data))
                
                for new_row, data in enumerate(rows_data):
                    # 全列のデータを設定
                    for col, cell_text in enumerate(data['cells']):
                        if col == 0:
                            # 訪問順序は新しい行番号+1に設定
                            item = QTableWidgetItem(str(new_row + 1))
                        else:
                            item = QTableWidgetItem(cell_text)
                        self.store_visits_table.setItem(new_row, col, item)
                    
                    # 星評価ウィジェットを再配置
                    star_widget = StarRatingWidget(
                        self.store_visits_table, 
                        rating=data['star_rating'], 
                        star_size=14
                    )
                    star_widget.rating_changed.connect(
                        lambda rating, r=new_row: self.on_star_rating_changed(r, rating)
                    )
                    self.store_visits_table.setCellWidget(new_row, 9, star_widget)
                
            finally:
                self.store_visits_table.blockSignals(False)
            
            # 移動時間を再計算
            self.recalc_travel_times()
            
            # 訪問順序を保存
            self.save_store_order()
            
            # 変更後の状態を保存
            self.save_table_state()
            
            # 計算結果を更新
            getattr(self, 'update_calculation_results', lambda: None)()
            
            QMessageBox.information(self, "完了", "訪問順序に基づいて行を並び替えました")
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"並び替え中にエラーが発生しました:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def add_store_visit_row(self):
        """店舗訪問行を追加"""
        # 変更前の状態を保存
        self.save_table_state()
        
        row = self.store_visits_table.rowCount()
        self.store_visits_table.insertRow(row)
        
        # 訪問順序を自動設定（編集可能）
        order_item = QTableWidgetItem(str(row + 1))
        self.store_visits_table.setItem(row, 0, order_item)
        
        # 星評価ウィジェットをセルに配置
        star_widget = StarRatingWidget(self.store_visits_table, rating=0, star_size=14)
        star_widget.rating_changed.connect(lambda rating, r=row: self.on_star_rating_changed(r, rating))
        self.store_visits_table.setCellWidget(row, 9, star_widget)
        
        # 変更後の状態を保存
        self.save_table_state()

    def on_star_rating_changed(self, row: int, rating: int):
        """星評価が変更されたときの処理"""
        # データ変更をUndoスタックに保存
        self.on_table_item_changed(None)

    def clear_all_rows(self):
        """すべての行をクリア"""
        reply = QMessageBox.question(
            self,
            "確認",
            "すべての行を削除しますか？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # 変更前の状態を保存
            self.save_table_state()
            
            self.store_visits_table.setRowCount(0)
            # 訪問順序を保存（クリア状態）
            self.save_store_order()
            getattr(self, 'update_calculation_results', lambda: None)()
            
            # 変更後の状態を保存
            self.save_table_state()
            
            QMessageBox.information(self, "完了", "すべての行を削除しました")

    def delete_store_visit_row(self):
        """選択された店舗訪問行を削除"""
        # 選択行の取得（行選択/セル選択の両対応）
        selected_indexes = self.store_visits_table.selectionModel().selectedRows()
        selected_rows = {idx.row() for idx in selected_indexes}
        if not selected_rows:
            current = self.store_visits_table.currentRow()
            if current >= 0:
                selected_rows = {current}
        if not selected_rows:
            QMessageBox.warning(self, "警告", "削除する行を選択してください")
            return
        
        # 変更前の状態を保存
        self.save_table_state()
        
        # 行番号を降順でソート（後ろから削除することでインデックスがずれない）
        for row in sorted(selected_rows, reverse=True):
            self.store_visits_table.removeRow(row)
        
        # 訪問順序を再設定
        self.update_visit_order()
        
        # 訪問順序を保存
        self.save_store_order()
        
        # 変更後の状態を保存
        self.save_table_state()
        
        getattr(self, 'update_calculation_results', lambda: None)()

    def save_data(self):
        """データを保存"""
        try:
            route_data = self.get_route_data()
            store_visits = self.get_store_visits_data()
            route_name_display = self.route_code_combo.currentText().strip() or route_data.get('route_code', '')
            
            if not route_data.get('route_code'):
                QMessageBox.warning(self, "警告", "ルートコードを入力してください")
                return
            
            self.save_store_order()
            
            stats = {}
            try:
                if self.calc_service:
                    stats = self.calc_service.calculate_route_statistics(route_data, store_visits) or {}
            except Exception:
                stats = {}
            stats_defaults = {
                'total_working_hours': 0,
                'estimated_hourly_rate': 0,
                'total_gross_profit': 0,
                'purchase_success_rate': 0,
                'avg_purchase_price': 0,
            }
            stats = {**stats_defaults, **stats}
            route_data.update(stats)
            route_data['route_name_display'] = route_name_display
            
            metrics = self.latest_summary_metrics or self._calculate_summary_metrics()
            if metrics:
                route_data['total_item_count'] = metrics.get('total_item_count', 0)
                route_data['total_gross_profit'] = metrics.get('total_gross_profit', 0.0)
                route_data['avg_purchase_price'] = metrics.get('avg_purchase_price', 0.0)
                route_data['total_purchase_amount'] = metrics.get('total_purchase_amount', 0.0)
                route_data['total_sales_amount'] = metrics.get('total_sales_amount', 0.0)
                route_data['total_working_hours'] = metrics.get('total_working_hours', 0.0)
                route_data['estimated_hourly_rate'] = metrics.get('estimated_hourly_rate', 0.0)
            else:
                def _int_safe(x):
                    try:
                        if x is None or x == '':
                            return 0
                        return int(float(x))
                    except Exception:
                        return 0
                def _float_safe(x):
                    try:
                        if x is None or x == '':
                            return 0.0
                        return float(x)
                    except Exception:
                        return 0.0
                route_data['total_item_count'] = sum(_int_safe(v.get('store_item_count')) for v in store_visits)
                route_data['total_gross_profit'] = sum(_float_safe(v.get('store_gross_profit')) for v in store_visits)
                route_data['total_purchase_amount'] = 0.0
                route_data['total_sales_amount'] = 0.0
                total_working_hours = self._calculate_total_working_hours(route_data.get('departure_time'), route_data.get('return_time'))
                route_data['total_working_hours'] = total_working_hours
                route_data['estimated_hourly_rate'] = self._calculate_hourly_rate(route_data.get('total_gross_profit'), total_working_hours)
            
            # 同日・同ルートが存在する場合は上書き
            if not self.current_route_id:
                existing = self.route_db.get_route_summary_by_date_code(route_data.get('route_date'), route_data.get('route_code'))
                if existing:
                    self.current_route_id = existing.get('id')
            
            if self.current_route_id:
                self.route_db.update_route_summary(self.current_route_id, route_data)
                existing_visits = self.route_db.get_store_visits_by_route(self.current_route_id)
                for visit in existing_visits:
                    self.route_db.delete_store_visit(visit['id'])
            else:
                self.current_route_id = self.route_db.add_route_summary(route_data)
            
            for visit in store_visits:
                visit['route_summary_id'] = self.current_route_id
                self.route_db.add_store_visit(visit)

            try:
                if self.route_visit_db:
                    self.route_visit_db.replace_route_visits(
                        route_data.get('route_date'),
                        route_data.get('route_code'),
                        route_name_display,
                        store_visits
                    )
            except Exception as db_err:
                logging.exception(f"ルート訪問履歴の保存に失敗しました: {db_err}")
            
            QMessageBox.information(self, "完了", "データを保存しました")
            self.data_saved.emit(self.current_route_id)
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"データ保存中にエラーが発生しました:\n{str(e)}")

    def get_route_data(self) -> Dict[str, Any]:
        """入力データを取得"""
        route_date = self.route_date_edit.dateTime().toString('yyyy-MM-dd')
        
        # route_dataから読み込んだ値を取得（テンプレート読み込み時に設定される）
        departure_time = self.route_data.get('departure_time') if hasattr(self, 'route_data') and self.route_data else None
        return_time = self.route_data.get('return_time') if hasattr(self, 'route_data') and self.route_data else None
        toll_fee_outbound = self.route_data.get('toll_fee_outbound') if hasattr(self, 'route_data') and self.route_data else 0
        toll_fee_return = self.route_data.get('toll_fee_return') if hasattr(self, 'route_data') and self.route_data else 0
        
        # デフォルト値の設定
        if not departure_time:
            departure_time = f"{route_date} 00:00:00"
        elif ' ' not in str(departure_time) and route_date:
            # HH:MM形式の場合はルート日付と結合
            departure_time = f"{route_date} {departure_time}:00"
        
        if not return_time:
            return_time = f"{route_date} 00:00:00"
        elif ' ' not in str(return_time) and route_date:
            # HH:MM形式の場合はルート日付と結合
            return_time = f"{route_date} {return_time}:00"
        
        return {
            'route_date': route_date,
            'route_code': self.get_selected_route_code(),
            'departure_time': departure_time,
            'return_time': return_time,
            'toll_fee_outbound': toll_fee_outbound if toll_fee_outbound else 0,
            'toll_fee_return': toll_fee_return if toll_fee_return else 0,
            'parking_fee': 0,
            'meal_cost': 0,
            'other_expenses': 0,
            'remarks': ''
        }

    def get_store_visits_data(self) -> List[Dict[str, Any]]:
        visits = []
        route_date = self.route_date_edit.dateTime().toString('yyyy-MM-dd')
        
        # 出発時刻・帰宅時間・往路高速代・復路高速代を除外する店舗コードリスト
        exclude_store_codes = ['出発時刻', '帰宅時刻', '往路高速代', '復路高速代']
        
        for i in range(self.store_visits_table.rowCount()):
            store_code = self._get_table_item(i, 1)
            
            # 出発時刻・帰宅時間・往路高速代・復路高速代は店舗訪問情報として扱わない
            if store_code in exclude_store_codes:
                continue
            
            star_widget = self.store_visits_table.cellWidget(i, 9)
            rating = star_widget.rating() if star_widget else 0
            
            # HH:MM形式の時間を取得してルート日付と結合
            in_time_str = self._get_table_item(i, 3)
            out_time_str = self._get_table_item(i, 4)
            
            # HH:MMを yyyy-MM-dd HH:MM:SS に変換
            store_in_time = self._combine_datetime(route_date, in_time_str)
            store_out_time = self._combine_datetime(route_date, out_time_str)
            
            visit = {
                'visit_order': len(visits) + 1,  # 除外した行を考慮して訪問順序を再計算
                'store_code': store_code,
                'store_name': self._get_table_item(i, 2),
                'store_in_time': store_in_time,
                'store_out_time': store_out_time,
                'stay_duration': self._safe_float(self._get_table_item(i, 5)),
                'travel_time_from_prev': self._safe_float(self._get_table_item(i, 6)),
                'store_gross_profit': self._safe_float(self._get_table_item(i, 7)),
                'store_item_count': self._safe_int(self._get_table_item(i, 8)),
                'store_rating': rating,
                'store_notes': self._get_table_item(i, 10)
            }
            visits.append(visit)
        return visits

    @staticmethod
    def _format_hm(value: Optional[str]) -> str:
        if not value:
            return ''
        text = str(value)
        if ' ' in text:
            text = text.split(' ')[1]
        return text[:5]

    def apply_route_snapshot(self, route_data: Dict[str, Any], visits: List[Dict[str, Any]]):
        """外部から受け取ったルート情報・店舗情報をUIに反映"""
        if route_data is None:
            route_data = {}
        date_str = route_data.get('route_date')
        if date_str:
            try:
                self.route_date_edit.setDateTime(QDateTime.fromString(date_str, 'yyyy-MM-dd'))
            except Exception:
                pass
        display_value = route_data.get('route_name_display') or route_data.get('route_code', '')
        if display_value:
            self.route_code_combo.blockSignals(True)
            try:
                self.update_route_codes()
                idx = self.route_code_combo.findText(display_value)
                if idx >= 0:
                    self.route_code_combo.setCurrentIndex(idx)
                else:
                    self.route_code_combo.addItem(display_value)
                    idx = self.route_code_combo.findText(display_value)
                    if idx >= 0:
                        self.route_code_combo.setCurrentIndex(idx)
                    else:
                        self.route_code_combo.setCurrentText(display_value)
            finally:
                self.route_code_combo.blockSignals(False)
        # 出発時間・帰宅時間・経費・備考は削除されたため、読み込み処理をスキップ
        
        self.store_visits_table.blockSignals(True)
        try:
            self.store_visits_table.setRowCount(len(visits))
            for i, visit in enumerate(visits):
                order_item = QTableWidgetItem(str(visit.get('visit_order', i + 1)))
                order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                self.store_visits_table.setItem(i, 0, order_item)
                self.store_visits_table.setItem(i, 1, QTableWidgetItem(visit.get('store_code', '')))
                self.store_visits_table.setItem(i, 2, QTableWidgetItem(visit.get('store_name', '')))
                self.store_visits_table.setItem(i, 3, QTableWidgetItem(self._format_hm(visit.get('store_in_time'))))
                self.store_visits_table.setItem(i, 4, QTableWidgetItem(self._format_hm(visit.get('store_out_time'))))
                self.store_visits_table.setItem(i, 5, QTableWidgetItem(str(visit.get('stay_duration', ''))))
                self.store_visits_table.setItem(i, 6, QTableWidgetItem(str(visit.get('travel_time_from_prev', ''))))
                self.store_visits_table.setItem(i, 7, QTableWidgetItem(str(visit.get('store_gross_profit', ''))))
                self.store_visits_table.setItem(i, 8, QTableWidgetItem(str(visit.get('store_item_count', ''))))
                rating = visit.get('store_rating', 0) or 0
                star_widget = StarRatingWidget(self.store_visits_table, rating=rating, star_size=14)
                self.store_visits_table.setCellWidget(i, 9, star_widget)
                self.store_visits_table.setItem(i, 10, QTableWidgetItem(visit.get('store_notes', '') or ''))
        finally:
            self.store_visits_table.blockSignals(False)
        self.update_visit_order()
        getattr(self, 'recalc_travel_times', lambda: None)()
        self.current_route_id = None
    
    def _combine_datetime(self, date_str: str, time_str: str) -> str:
        """ルート日付と時間（HH:MM）を結合してDATETIME形式にする"""
        if not time_str or not time_str.strip():
            return ''
        
        try:
            # HH:MM形式をチェック
            parts = time_str.strip().split(':')
            if len(parts) != 2:
                return ''
            
            h = int(parts[0])
            m = int(parts[1])
            
            if not (0 <= h <= 23 and 0 <= m <= 59):
                return ''
            
            # yyyy-MM-dd HH:MM:SS形式で返す
            return f"{date_str} {h:02d}:{m:02d}:00"
        except (ValueError, TypeError):
            return ''

    def _calculate_total_working_hours(self, departure_time: Optional[str], return_time: Optional[str]) -> float:
        """総稼働時間（時間）を計算"""
        dep_dt = None
        ret_dt = None
        if self.calc_service:
            dep_dt = CalculationService.parse_datetime_string(departure_time)
            ret_dt = CalculationService.parse_datetime_string(return_time)
        if not dep_dt and departure_time:
            try:
                dep_dt = datetime.fromisoformat(departure_time.replace('Z', '+00:00'))
            except Exception:
                dep_dt = None
        if not ret_dt and return_time:
            try:
                ret_dt = datetime.fromisoformat(return_time.replace('Z', '+00:00'))
            except Exception:
                ret_dt = None
        if dep_dt and ret_dt:
            if self.calc_service:
                hours = CalculationService.calculate_total_working_hours(dep_dt, ret_dt)
                if hours is not None:
                    return max(round(hours, 2), 0.0)
            diff = (ret_dt - dep_dt).total_seconds() / 3600.0
            if diff < 0:
                diff = 0.0
            return round(diff, 2)
        return 0.0

    def _calculate_hourly_rate(self, gross_profit: Optional[float], working_hours: Optional[float]) -> float:
        """想定時給を計算"""
        gross = gross_profit or 0.0
        hours = working_hours or 0.0
        if self.calc_service:
            value = CalculationService.calculate_hourly_rate(gross, hours)
            if value is not None:
                return max(round(value, 2), 0.0)
        if hours > 0:
            return round(gross / hours, 2)
        return 0.0

    def _calculate_summary_metrics(self) -> Optional[Dict[str, Any]]:
        """現在の入力値からサマリー指標を算出"""
        try:
            route_date = self.route_date_edit.dateTime().toString('yyyy-MM-dd')
        except Exception:
            route_date = ''

        route_code = self.get_selected_route_code()
        route_name_display = self.route_code_combo.currentText().strip() or route_code
        try:
            if route_code:
                resolved = self.store_db.get_route_name_by_code(route_code)
                if resolved:
                    route_name_display = resolved
        except Exception:
            pass

        def to_float(value) -> Optional[float]:
            if value is None:
                return None
            try:
                if isinstance(value, str):
                    value = value.replace(',', '').strip()
                    if value == '':
                        return None
                return float(value)
            except (ValueError, TypeError):
                return None

        def to_int(value) -> Optional[int]:
            val = to_float(value)
            if val is None:
                return None
            return int(round(val))

        inventory_df = None
        if self.inventory_widget and hasattr(self.inventory_widget, 'inventory_data'):
            try:
                data = self.inventory_widget.inventory_data
                if data is not None:
                    if hasattr(data, 'iterrows'):
                        inventory_df = data.copy()
                    elif isinstance(data, list):
                        inventory_df = pd.DataFrame(data)
            except Exception:
                inventory_df = None

        total_purchase = 0.0
        total_sales = 0.0
        total_profit = 0.0
        total_items = 0

        if inventory_df is not None and len(inventory_df) > 0:
            for _, row in inventory_df.iterrows():
                qty = to_float(
                    row.get('仕入れ個数')
                    or row.get('purchase_count')
                    or row.get('quantity')
                    or row.get('quantityPurchased')
                    or row.get('数量')
                ) or 0.0
                purchase_price = to_float(row.get('仕入れ価格') or row.get('purchasePrice') or row.get('cost_price'))
                sale_price = to_float(row.get('販売予定価格') or row.get('plannedPrice') or row.get('expected_sale_price'))
                profit_unit = to_float(row.get('見込み利益') or row.get('expected_profit') or row.get('profit') or row.get('expectedProfit'))

                total_items += int(round(qty))
                if purchase_price is not None:
                    total_purchase += purchase_price * qty
                if sale_price is not None:
                    total_sales += sale_price * qty
                if profit_unit is not None:
                    total_profit += profit_unit * qty

            if total_sales == 0 and total_purchase and total_profit:
                total_sales = total_purchase + total_profit
            if total_profit == 0 and total_sales and total_purchase:
                total_profit = total_sales - total_purchase

        table_items = 0
        table_profit = 0.0
        for row in range(self.store_visits_table.rowCount()):
            qty = to_float(self._get_table_item(row, 8)) or 0.0
            table_items += int(round(qty))
            profit_val = to_float(self._get_table_item(row, 7)) or 0.0
            table_profit += profit_val

        if total_items == 0:
            total_items = table_items
        if total_profit == 0:
            total_profit = table_profit
        if total_sales == 0 and total_purchase and total_profit:
            total_sales = total_purchase + total_profit

        avg_purchase_price = 0.0
        if total_items > 0 and total_purchase:
            avg_purchase_price = total_purchase / total_items

        route_data_preview = self.get_route_data()
        working_hours = self._calculate_total_working_hours(
            route_data_preview.get('departure_time'),
            route_data_preview.get('return_time')
        )
        hourly_rate = self._calculate_hourly_rate(total_profit, working_hours)

        return {
            'route_date': route_date,
            'route_code': route_code,
            'route_name': route_name_display,
            'total_item_count': total_items,
            'total_purchase_amount': total_purchase,
            'total_sales_amount': total_sales,
            'total_gross_profit': total_profit,
            'avg_purchase_price': avg_purchase_price,
            'total_working_hours': working_hours,
            'estimated_hourly_rate': hourly_rate,
        }

    def _get_table_item(self, row: int, col: int) -> str:
        item = self.store_visits_table.item(row, col)
        return item.text() if item else ''

    def _safe_float(self, value: str) -> Optional[float]:
        try:
            return float(value) if value else None
        except (ValueError, TypeError):
            return None

    def _safe_int(self, value: str) -> Optional[int]:
        try:
            return int(float(value)) if value else None
        except (ValueError, TypeError):
            return None
    
    def recalculate_matching(self):
        """照合再計算処理: 仕入管理タブのデータから想定粗利・仕入れ点数を再計算"""
        try:
            if not self.current_route_id:
                QMessageBox.warning(self, "警告", "先にルートを保存してください")
                return
            
            # 仕入管理データの確認
            if not self.inventory_widget:
                QMessageBox.warning(self, "警告", "仕入管理ウィジェットへの参照がありません")
                return
            
            # 🔥 重要: テーブルから最新のデータを再取得（手入力データを含む）
            # テーブルの内容をinventory_dataに同期
            if hasattr(self.inventory_widget, 'sync_inventory_data_from_table'):
                sync_success = self.inventory_widget.sync_inventory_data_from_table()
                if not sync_success:
                    QMessageBox.warning(self, "警告", "テーブルデータの取得に失敗しました")
                    return
            
            # テーブルから直接データを取得（より確実な方法）
            if hasattr(self.inventory_widget, 'get_table_data'):
                table_data = self.inventory_widget.get_table_data()
                if table_data is not None and len(table_data) > 0:
                    inventory_data = table_data
                    print(f"\n=== 照合再計算: テーブルからデータ取得 ===")
                    print(f"取得件数: {len(inventory_data)}")
                    # デバッグ: K1-010の件数を確認
                    k1_010_count = 0
                    if '仕入先' in inventory_data.columns:
                        k1_010_count = len(inventory_data[inventory_data['仕入先'].astype(str).str.strip().str.replace('(', '').str.replace(')', '') == 'K1-010'])
                    print(f"K1-010の件数（テーブルから取得）: {k1_010_count}")
                    # inventory_dataも更新しておく（今後の処理で使用される可能性があるため）
                    self.inventory_widget.inventory_data = table_data.copy()
                    self.inventory_widget.filtered_data = table_data.copy()
                else:
                    inventory_data = self.inventory_widget.inventory_data
                    print(f"\n=== 照合再計算: テーブルデータが空のため、既存のinventory_dataを使用 ===")
            else:
                inventory_data = self.inventory_widget.inventory_data
                print(f"\n=== 照合再計算: get_table_dataメソッドがないため、既存のinventory_dataを使用 ===")
            
            if inventory_data is None or len(inventory_data) == 0:
                QMessageBox.warning(self, "警告", "仕入管理にデータがありません")
                return
            
            # 確認ダイアログ
            reply = QMessageBox.question(
                self,
                "照合再計算",
                "想定粗利・仕入れ点数を再計算しますか？\n\n他の項目は変更されません。",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
            
            # データをJSON形式に変換（NaN値を事前に処理）
            clean_data = inventory_data.fillna('')
            purchase_data = clean_data.to_dict(orient="records")
            
            # 粗利再計算
            self._update_route_gross_profit_from_inventory(purchase_data)
            self.update_calculation_results()
            
            QMessageBox.information(self, "完了", "照合再計算が完了しました")
            
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"照合再計算中にエラーが発生しました:\n{str(e)}")
    
    def _update_route_gross_profit_from_inventory(self, inventory_data: List[Dict[str, Any]]):
        """
        仕入管理データから店舗コード別の粗利を集計してルートサマリーを更新
        
        計算方法:
        - 想定粗利: 各商品の「仕入れ個数 × 見込み利益」を店舗別に合計
        - 仕入れ点数: 店舗毎の「仕入れ個数」の総数（SUM）
        """
        try:
            if not self.current_route_id:
                return
            
            # 店舗コード別の粗利と仕入れ個数を集計
            store_profits = {}  # 想定粗利の合計（仕入れ個数 × 見込み利益）
            store_item_counts = {}  # 仕入れ個数の総数
            print(f"\n照合再計算開始: inventory_data件数={len(inventory_data)}")
            # デバッグ: K1-010のデータを全て確認
            k1_010_items = []
            for idx, item in enumerate(inventory_data):
                store_code_raw = item.get('仕入先') or item.get('supplier')
                if store_code_raw:
                    store_code_cleaned = store_code_raw.strip().strip('()') if isinstance(store_code_raw, str) else str(store_code_raw).strip().strip('()')
                    if store_code_cleaned == 'K1-010':
                        k1_010_items.append((
                            idx, 
                            item.get('商品名', 'N/A')[:30], 
                            item.get('仕入れ個数'), 
                            item.get('見込み利益')
                        ))
            
            if k1_010_items:
                print(f"  K1-010のデータ一覧 ({len(k1_010_items)}件):")
                for idx, name, count, profit in k1_010_items:
                    print(f"    行{idx}: {name}, 仕入れ個数={count}, 見込み利益={profit}")
            
            for idx, item in enumerate(inventory_data):
                store_code = item.get('仕入先') or item.get('supplier')
                if not store_code:
                    continue
                
                # 括弧を取り除いて正規化（例: "(K1-010)" → "K1-010"）
                if isinstance(store_code, str):
                    store_code = store_code.strip().strip('()')
                
                # 仕入れ個数と見込み利益を取得
                item_count = self._safe_float(item.get('仕入れ個数') or item.get('purchase_count') or item.get('quantity'))
                expected_profit = self._safe_float(item.get('見込み利益') or item.get('expected_profit') or item.get('profit'))
                
                # デバッグ: 値がNoneのデータを確認
                if store_code == 'K1-010':
                    if item_count is None or expected_profit is None:
                        print(f"  K1-010で値None検出: 仕入れ個数={item.get('仕入れ個数')}, 見込み利益={item.get('見込み利益')}")
                
                # 初期化
                if store_code not in store_profits:
                    store_profits[store_code] = 0
                if store_code not in store_item_counts:
                    store_item_counts[store_code] = 0
                
                # 仕入れ個数の総数を計算（店舗毎の合計）
                if item_count is not None and item_count >= 0:
                    store_item_counts[store_code] += int(item_count)
                
                # 想定粗利の計算（仕入れ個数 × 見込み利益）
                if item_count is not None and expected_profit is not None:
                    # 仕入れ個数が0以上の場合は計算に含める
                    if item_count >= 0:
                        profit_per_store = item_count * expected_profit
                        store_profits[store_code] += profit_per_store
                        # デバッグ: K1-010を含むデータを出力
                        if store_code == 'K1-010':
                            print(f"  K1-010処理: 仕入れ個数={item_count}, 見込み利益={expected_profit}, 粗利={profit_per_store}")
            
            # デバッグ: 集計結果を出力
            print(f"\n=== 照合再計算: 店舗別粗利集計結果 ===")
            for code in store_profits.keys():
                profit = store_profits.get(code, 0)
                item_count = store_item_counts.get(code, 0)
                print(f"  {code}: 粗利={profit}円, 仕入れ個数={item_count}")
            
            # ルートサマリーの店舗訪問詳細を取得
            visits = self.route_db.get_store_visits_by_route(self.current_route_id)
            
            # 各店舗訪問に粗利を設定
            for visit in visits:
                store_code = visit.get('store_code')
                if store_code in store_profits:
                    # 整数に変換（小数点なし）
                    visit['store_gross_profit'] = int(store_profits[store_code])
                    # 仕入れ点数は店舗毎の仕入れ個数の総数
                    visit['store_item_count'] = store_item_counts.get(store_code, 0)
                    # 粗利が0より大きい場合は仕入れ成功とみなす
                    visit['purchase_success'] = (store_profits[store_code] > 0)
                    print(f"  → {store_code} を更新: 粗利={visit['store_gross_profit']}, 仕入れ個数={visit['store_item_count']}")
                    logging.info(f"照合再計算: {store_code} を更新 - 粗利: {visit['store_gross_profit']}, 仕入れ個数: {visit['store_item_count']}")
                else:
                    # マッチしない店舗は0に設定
                    visit['store_gross_profit'] = 0
                    visit['store_item_count'] = 0
                    visit['purchase_success'] = False
                    print(f"  → {store_code} はマッチなし")
                    logging.info(f"照合再計算: {store_code} はマッチなし")
                
                # 店舗訪問詳細を更新
                result = self.route_db.update_store_visit(visit['id'], visit)
                if not result:
                    print(f"  ⚠️ {store_code} のDB更新に失敗")
                    logging.warning(f"照合再計算: {store_code} のDB更新に失敗")
            
            # テーブルを更新
            if self.current_route_id:
                # デバッグ: 更新前のDB状態を確認
                print(f"\n照合再計算完了: 更新した店舗数={len([v for v in visits if v.get('store_code') in store_profits])}")
                self.load_saved_data(self.current_route_id)

                # 画面のテーブルにも即時反映（DB読込に失敗した場合のフォールバック）
                try:
                    code_col = 1  # 店舗コード列
                    gross_col = 7  # 想定粗利列
                    count_col = 8  # 仕入れ点数列
                    for r in range(self.store_visits_table.rowCount()):
                        code_item = self.store_visits_table.item(r, code_col)
                        if not code_item:
                            continue
                        code = (code_item.text() or '').strip()
                        if not code:
                            continue
                        if code in store_profits or code in store_item_counts:
                            gp = int(store_profits.get(code, 0))
                            cnt = int(store_item_counts.get(code, 0))
                            self.store_visits_table.setItem(r, gross_col, QTableWidgetItem(str(gp)))
                            self.store_visits_table.setItem(r, count_col, QTableWidgetItem(str(cnt)))
                except Exception as _e:
                    print(f"UI反映フォールバックでエラー: {_e}")
            
            self.auto_calculate_store_ratings()

        except Exception as e:
            # エラーはログに記録するが、ユーザーには通知しない（主要処理は成功しているため）
            logging.error(f"ルート粗利更新エラー: {str(e)}")

class StoreSelectDialog(QDialog):
    """店舗マスタ一覧を表示して選択させるダイアログ"""
    def __init__(self, store_db: StoreDatabase, parent=None):
        super().__init__(parent)
        self.setWindowTitle("店舗マスタから追加")
        self.resize(700, 480)
        self.store_db = store_db
        self.selected_rows: List[Dict[str, Any]] = []

        layout = QVBoxLayout(self)
        search_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("コード/店舗名で検索")
        search_btn = QPushButton("検索")
        search_btn.clicked.connect(self.filter_rows)
        search_layout.addWidget(self.search_edit)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["ルート名", "店舗コード", "店舗名"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.MultiSelection)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        layout.addWidget(self.table)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

        self.load_rows()

    def load_rows(self):
        try:
            stores = self.store_db.list_stores()
        except Exception:
            stores = []
        self.all_rows = stores
        self.populate(stores)

    def populate(self, stores: List[Dict[str, Any]]):
        self.table.setRowCount(len(stores))
        for i, s in enumerate(stores):
            self.table.setItem(i, 0, QTableWidgetItem(str(s.get('affiliated_route_name', ''))))
            self.table.setItem(i, 1, QTableWidgetItem(str(s.get('supplier_code', ''))))
            self.table.setItem(i, 2, QTableWidgetItem(str(s.get('store_name', ''))))

    def filter_rows(self):
        q = self.search_edit.text().strip().lower()
        if not q:
            self.populate(self.all_rows)
            return
        filtered = [s for s in self.all_rows if q in str(s.get('supplier_code','')).lower() or q in str(s.get('store_name','')).lower()]
        self.populate(filtered)

    def get_selected_stores(self) -> List[Dict[str, Any]]:
        rows = []
        for idx in self.table.selectionModel().selectedRows():
            r = idx.row()
            code = self.table.item(r, 1).text() if self.table.item(r, 1) else ''
            name = self.table.item(r, 2).text() if self.table.item(r, 2) else ''
            route = self.table.item(r, 0).text() if self.table.item(r, 0) else ''
            rows.append({'supplier_code': code, 'store_name': name, 'affiliated_route_name': route})
        return rows


class SavedRoutesDialog(QDialog):
    """保存済みルートの一覧から選択して読込/削除するダイアログ"""
    def __init__(self, route_db: RouteDatabase, store_db: StoreDatabase, parent=None):
        super().__init__(parent)
        self.setWindowTitle("保存履歴")
        self.resize(720, 420)
        self.route_db = route_db
        self.store_db = store_db
        self._result_action = None  # 'load' or 'delete'
        self._result_id = None

        layout = QVBoxLayout(self)

        # フィルタ行（デフォルトは全件表示。必要時のみチェックして絞り込む）
        filt = QHBoxLayout()
        from PySide6.QtWidgets import QDateEdit
        self.chk_date = QCheckBox("仕入れ日")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDateTime(QDateTime.currentDateTime())
        self.chk_route = QCheckBox("ルート名")
        self.route_combo = QComboBox()
        self.route_combo.setEditable(True)
        try:
            names = self.store_db.get_route_names()
            for n in names:
                self.route_combo.addItem(n)
        except Exception:
            pass
        search_btn = QPushButton("検索")
        search_btn.clicked.connect(self._reload)
        filt.addWidget(self.chk_date)
        filt.addWidget(self.date_edit)
        filt.addWidget(self.chk_route)
        filt.addWidget(self.route_combo)
        filt.addWidget(search_btn)
        layout.addLayout(filt)

        # 一覧テーブル
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["ID", "仕入れ日", "ルートコード（ルート名）", "最終更新"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        # 列幅を内容に合わせて自動調整（最終更新列が切れないように）
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        layout.addWidget(self.table)

        # ボタン
        btns = QDialogButtonBox()
        self.load_btn = QPushButton("読み込み")
        self.del_btn = QPushButton("削除")
        self.cancel_btn = QPushButton("閉じる")
        btns.addButton(self.load_btn, QDialogButtonBox.AcceptRole)
        btns.addButton(self.del_btn, QDialogButtonBox.ActionRole)
        btns.addButton(self.cancel_btn, QDialogButtonBox.RejectRole)
        layout.addWidget(btns)

        self.load_btn.clicked.connect(self._on_load)
        self.del_btn.clicked.connect(self._on_delete)
        self.cancel_btn.clicked.connect(self.reject)

        self._reload()

    def _reload(self):
        try:
            route_name = self.route_combo.currentText().strip()
            route_code = None
            if self.chk_route.isChecked() and route_name:
                route_code = self.store_db.get_route_code_by_name(route_name)
        except Exception:
            route_code = None
        if self.chk_date.isChecked():
            day = self.date_edit.date().toString('yyyy-MM-dd')
            start = day; end = day
        else:
            start = None; end = None
        rows = self.route_db.list_route_summaries(start_date=start, end_date=end, route_code=route_code)
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            rid = str(r.get('id',''))
            rdate = str(r.get('route_date',''))
            rcode = str(r.get('route_code',''))
            # ルート名に変換
            try:
                rname = self.store_db.get_route_name_by_code(rcode) or ''
            except Exception:
                rname = ''
            code_display = f"{rcode}"
            if rname:
                code_display = f"{rcode}（{rname}）"
            updated_at = str(r.get('updated_at',''))

            self.table.setItem(i, 0, QTableWidgetItem(rid))
            self.table.setItem(i, 1, QTableWidgetItem(rdate))
            self.table.setItem(i, 2, QTableWidgetItem(code_display))
            self.table.setItem(i, 3, QTableWidgetItem(updated_at))
        # データ投入後に列幅を最適化
        try:
            self.table.resizeColumnsToContents()
        except Exception:
            pass

    def _selected_id(self) -> Optional[int]:
        sel = self.table.selectionModel().selectedRows() if self.table.selectionModel() else []
        if not sel:
            return None
        r = sel[0].row()
        item = self.table.item(r, 0)
        try:
            return int(item.text()) if item else None
        except Exception:
            return None

    def _on_load(self):
        rid = self._selected_id()
        if rid is None:
            QMessageBox.information(self, "情報", "読み込む行を選択してください")
            return
        self._result_action = 'load'
        self._result_id = rid
        self.accept()

    def _on_delete(self):
        rid = self._selected_id()
        if rid is None:
            QMessageBox.information(self, "情報", "削除する行を選択してください")
            return
        self._result_action = 'delete'
        self._result_id = rid
        self.accept()

    def get_result(self):
        return self._result_action, self._result_id


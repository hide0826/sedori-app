#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AmazonテンプレートExcelファイルの「テンプレート」シートを読み込む
"""
import sys
import os
from pathlib import Path

# パスを追加
sys.path.insert(0, str(Path(__file__).parent))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)

def analyze_template():
    file_path = Path(__file__).parent / "ListingLoader.xlsm"
    
    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        return
    
    print(f"Reading file: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        print(f"✓ Workbook loaded successfully")
        print(f"\nAvailable sheets: {wb.sheetnames}")
        
        if "テンプレート" not in wb.sheetnames:
            print(f"\nERROR: 'テンプレート' sheet not found!")
            print(f"Available sheets: {wb.sheetnames}")
            return
        
        ws = wb["テンプレート"]
        print(f"\n✓ 'テンプレート' sheet found")
        print(f"  Max row: {ws.max_row}")
        print(f"  Max column: {ws.max_column}")
        
        # ヘッダー行を取得
        print(f"\n【ヘッダー行（1行目）】")
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                headers.append((col_idx, str(cell_value)))
                print(f"  列{col_idx}: {cell_value}")
        
        # サンプルデータ（2-5行目）
        print(f"\n【データ行のサンプル（2-5行目）】")
        for row_idx in range(2, min(6, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):  # 最初の20列まで
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(str(cell_value) if cell_value is not None else "")
            print(f"  行{row_idx}: {row_values}")
        
        # 結果をファイルに保存
        output_file = Path(__file__).parent / "template_analysis.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("=== Amazon Template Excel Analysis ===\n\n")
            f.write(f"File: {file_path}\n")
            f.write(f"Sheets: {wb.sheetnames}\n\n")
            f.write(f"Template Sheet:\n")
            f.write(f"  Max row: {ws.max_row}\n")
            f.write(f"  Max column: {ws.max_column}\n\n")
            f.write("Headers:\n")
            for col_idx, header in headers:
                f.write(f"  Column {col_idx}: {header}\n")
            f.write("\nSample data:\n")
            for row_idx in range(2, min(6, ws.max_row + 1)):
                row_values = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_values.append(str(cell_value) if cell_value is not None else "")
                f.write(f"  Row {row_idx}: {row_values}\n")
        
        print(f"\n✓ Results saved to: {output_file}")
        
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_template()




